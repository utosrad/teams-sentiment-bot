"""
Interac Sentiment Analysis Bot
- Scrapes Reddit, X, RedFlagDeals, news for Interac mentions 4x/day
- Splits people vs press signals
- Alerts on sentiment drops
- Configurable via prompts.json
"""

import os
import json
import logging
import smtplib
import asyncio
import re
import html
from time import monotonic
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from collections import defaultdict
from urllib.parse import urlparse, urlunparse, quote
from email.utils import parsedate_to_datetime
import xml.etree.ElementTree as ET

import httpx
from ddgs import DDGS
from telegram import Update
from telegram.error import BadRequest
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ─── Config ───────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
KIMI_API_KEY = os.environ["KIMI_API_KEY"]
KIMI_API_URL = os.environ.get("KIMI_API_URL", "https://api.moonshot.ai/v1/chat/completions")
KIMI_MODEL = os.environ.get("KIMI_MODEL", "kimi-k2.5-preview")
PORT = int(os.environ.get("PORT", 3978))
WEBHOOK_URL = os.environ.get("WEBHOOK_URL", "")
ADMIN_IDS = {int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x}
DAILY_LIMIT = int(os.environ.get("DAILY_LIMIT", "5"))

EMAIL_ENABLED = os.environ.get("EMAIL_ENABLED", "0") == "1"
EMAIL_SEND_MODE = os.environ.get("EMAIL_SEND_MODE", "alert").lower()
EMAIL_ALERT_DEDUP = os.environ.get("EMAIL_ALERT_DEDUP", "1") == "1"
EMAIL_COOLDOWN_MINUTES = int(os.environ.get("EMAIL_COOLDOWN_MINUTES", "0"))
EMAIL_WEEKLY_DAY = os.environ.get("EMAIL_WEEKLY_DAY", "monday").strip().lower()
EMAIL_WEEKLY_HOUR = int(os.environ.get("EMAIL_WEEKLY_HOUR", "9"))
EMAIL_PROVIDER = os.environ.get("EMAIL_PROVIDER", "smtp").strip().lower()

SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
RESEND_API_URL = os.environ.get("RESEND_API_URL", "https://api.resend.com/emails")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "")
EMAIL_TO = [x.strip() for x in os.environ.get("EMAIL_TO", "").split(",") if x.strip()]
EMAIL_SUBJECT_PREFIX = os.environ.get("EMAIL_SUBJECT_PREFIX", "Interac Intelligence")
SOURCE_LEDGER_PUBLIC_URL = os.environ.get("SOURCE_LEDGER_PUBLIC_URL", "").strip()
MAX_MENTION_AGE_DAYS = int(os.environ.get("MAX_MENTION_AGE_DAYS", "120"))
QUALITY_STRICT = os.environ.get("QUALITY_STRICT", "1") == "1"
TWITTERAPI_IO_KEY = os.environ.get("TWITTERAPI_IO_KEY", "")
TWITTERAPI_IO_URL = "https://api.twitterapi.io/twitter/tweet/advanced_search"

EASTERN_TZ = ZoneInfo("America/Toronto")
EST = EASTERN_TZ  # America/Toronto — used for scheduling and rate-limit day boundaries

subscribed_chats: set[int] = set()
last_report: str = ""
last_mentions_raw: str = ""
# Canonical URL (see _canonical_url_for_date_lookup) -> "Month DD, YYYY" from last biweekly fetch
last_biweekly_url_dates: dict[str, str] = {}
last_quarterly_url_dates: dict[str, str] = {}
last_email_sent_at: datetime | None = None
last_weekly_email_key: str | None = None
last_quarterly_email_key: str | None = None

# Per-user daily rate limiting
user_usage: dict[int, dict] = defaultdict(lambda: {"count": 0, "date": None})
active_tasks: set[asyncio.Task] = set()


def _track_current_task() -> asyncio.Task | None:
    task = asyncio.current_task()
    if task is not None:
        active_tasks.add(task)
    return task


def _untrack_task(task: asyncio.Task | None) -> None:
    if task is not None:
        active_tasks.discard(task)


def _cancel_active_tasks(*, exclude: asyncio.Task | None = None) -> int:
    cancelled = 0
    for task in list(active_tasks):
        if exclude is not None and task is exclude:
            continue
        if task.done():
            active_tasks.discard(task)
            continue
        task.cancel()
        cancelled += 1
    return cancelled


def now_est() -> str:
    now = datetime.now(EASTERN_TZ)
    abbr = "EDT" if now.dst() else "EST"
    return now.strftime(f"%Y-%m-%d %I:%M %p {abbr}")


def check_rate_limit(user_id: int) -> tuple[bool, int]:
    if user_id in ADMIN_IDS:
        return True, -1

    today = datetime.now(EST).date()
    usage = user_usage[user_id]

    if usage["date"] != today:
        usage["count"] = 0
        usage["date"] = today

    if usage["count"] >= DAILY_LIMIT:
        return False, 0

    usage["count"] += 1
    return True, DAILY_LIMIT - usage["count"]


# ─── Prompt Config ────────────────────────────────────────────────────────────
def load_prompts() -> dict:
    base_dir = Path(__file__).parent
    config_path = base_dir / "prompts.json"
    with open(config_path) as f:
        config = json.load(f)

    prompt_files = config.get("prompt_files", {})
    default_prompt_files = {
        "followup_prompt": "prompts/followup_prompt.md",
    }
    for prompt_key, default_path in default_prompt_files.items():
        rel_path = prompt_files.get(prompt_key, default_path)
        prompt_path = base_dir / rel_path
        if prompt_path.exists():
            config[prompt_key] = prompt_path.read_text().strip()
        elif prompt_key not in config:
            raise FileNotFoundError(f"Missing prompt file for {prompt_key}: {prompt_path}")

    # Optional extra prompts (e.g. historical_prompt).
    for prompt_key, rel_path in prompt_files.items():
        if prompt_key in config:
            continue
        if not prompt_key.endswith("_prompt"):
            continue
        prompt_path = base_dir / rel_path
        if prompt_path.exists():
            config[prompt_key] = prompt_path.read_text().strip()
        else:
            raise FileNotFoundError(f"Missing prompt file for {prompt_key}: {prompt_path}")

    ledger_url = source_ledger_display_url()
    for k, v in list(config.items()):
        if isinstance(v, str) and "{source_ledger_url}" in v:
            config[k] = v.replace("{source_ledger_url}", ledger_url)

    return config


def source_ledger_display_url() -> str:
    """Human-facing location of the per-source Excel ledger (env or on-server fallback)."""
    if SOURCE_LEDGER_PUBLIC_URL:
        return SOURCE_LEDGER_PUBLIC_URL
    return "On server: state/source_ledger.xlsx (not web-hosted)"


def _source_ledger_footer() -> str:
    return f"\n\nSource ledger: {source_ledger_display_url()}"


def _strip_model_ledger_lines(text: str) -> str:
    """Remove trailing `Source ledger: ...` lines models may emit; app adds one canonical footer."""
    lines = (text or "").strip().splitlines()
    while lines and "source ledger:" in lines[-1].lower():
        lines.pop()
    return "\n".join(lines).strip()


# ─── Web Scraping ─────────────────────────────────────────────────────────────
def lookback_hours_to_tbs(lookback_hours: int) -> str:
    # Google-style time filters used by query config.
    if lookback_hours <= 24:
        return "qdr:d"
    if lookback_hours <= 24 * 7:
        return "qdr:w"
    return "qdr:m"


def normalize_tbs(tbs: str) -> str:
    """Normalize tbs value. Empty string means 'all time' (no time filter)."""
    if tbs in ("", "all"):
        return ""
    supported = {"qdr:d", "qdr:w", "qdr:m", "qdr:y"}
    return tbs if tbs in supported else "qdr:m"


def _has_site_restriction(query: str) -> bool:
    return "site:" in query.lower()


_search_errors: list[str] = []

# ── Spam / low-quality domain blocklist ───────────────────────────────────────
# Blocks casino affiliates, gambling sites, and known spam domains that rank
# for "e-Transfer" because Canadians use it for online gambling deposits.
_BLOCKED_DOMAINS: set[str] = {
    "bodog.com", "bovada.lv", "betway.com", "888casino.com", "jackpot.com",
    "casumo.com", "casinorewards.ca", "enmarie.com", "mesnmw.org",
    "spinpalace.com", "playojo.ca", "betmgm.ca", "draftkings.com",
    "fanduel.com", "pointsbet.ca", "proline.ca", "bet365.com",
}
_BLOCKED_DOMAIN_KEYWORDS: tuple[str, ...] = (
    "casino", "gambling", "betting", "poker", "slots", "wagering",
    "sportsbook", "sportsbetting", "onlinegambling",
)


def _is_blocked_domain(url: str) -> bool:
    """Return True if the URL belongs to a spam/casino/gambling domain."""
    try:
        domain = re.sub(r"^www\.", "", urlparse(url).netloc.lower())
    except Exception:
        return False
    if domain in _BLOCKED_DOMAINS:
        return True
    return any(kw in domain for kw in _BLOCKED_DOMAIN_KEYWORDS)


# ── Market pulse quality filters ──────────────────────────────────────────────
# Low-quality SEO blogs that explain what products are rather than reporting
# on market developments. Filtered from press and competitor buckets only.
_MARKET_LOW_QUALITY_DOMAINS: set[str] = {
    "moonrockcanada.net", "newcomersetup.ca", "briefglance.com",
    "howtosavemoney.ca", "frugalflyer.ca", "milesopedia.com",
    "mybanktracker.com", "savvynewcanadians.com", "hardbacon.ca",
}

# Explainer content — generic "what is X" SEO articles, not market news.
_EXPLAINER_RE = re.compile(
    r"(what is (interac|e-transfer|e transfer|wise|paypal|wealthsimple)|"
    r"safe and secure way to (send|transfer)|"
    r"send money (electronically|instantly|directly) (to|from)|"
    r"no need for (cash|cheques|checks)|"
    r"how (to|does) (send|use|work).{0,30}(e-transfer|interac|etransfer))",
    re.IGNORECASE,
)


def _is_low_quality_market_content(mention: dict) -> bool:
    """Return True for SEO explainer content that has no market intelligence value."""
    url = mention.get("link") or ""
    try:
        domain = re.sub(r"^www\.", "", urlparse(url).netloc.lower())
    except Exception:
        domain = ""
    if domain in _MARKET_LOW_QUALITY_DOMAINS:
        return True
    text = f"{mention.get('title', '')} {mention.get('snippet', '')}".strip()
    return bool(_EXPLAINER_RE.search(text))


def _classify_channel_and_source(link: str) -> tuple[str, str]:
    url = (link or "").lower()
    if "reddit.com" in url:
        return "people", "Reddit"
    if "x.com" in url or "twitter.com" in url:
        return "people", "X/Twitter"
    if "redflagdeals.com" in url:
        return "people", "RedFlagDeals"
    if "forum" in url or "community" in url:
        return "people", "Forum"
    return "press", "News/Other"


def _source_quality_tier(link: str, channel: str) -> str:
    url = (link or "").lower()
    if channel == "people":
        return "tier1_user_generated"
    if any(d in url for d in ["reddit.com", "x.com", "twitter.com", "redflagdeals.com"]):
        return "tier1_user_generated"
    if any(
        d in url
        for d in [
            "reuters.com",
            "bloomberg.com",
            "cbc.ca",
            "theglobeandmail.com",
            "financialpost.com",
        ]
    ):
        return "tier2_reported"
    return "tier3_commentary_or_unknown"


def _detect_brands(text: str) -> str:
    content = (text or "").lower()
    brand_order = [
        ("interac", "Interac"),
        ("wise", "Wise"),
        ("paypal", "PayPal"),
        ("apple pay", "ApplePay"),
        ("google pay", "GooglePay"),
        ("samsung pay", "SamsungPay"),
        ("venmo", "Venmo"),
        ("cash app", "CashApp"),
    ]
    found = [name for token, name in brand_order if token in content]
    return ", ".join(found) if found else "Unknown"


def _detect_use_case(text: str) -> str:
    content = (text or "").lower()
    if any(x in content for x in ["cross-border", "outside canada", "international", "remittance"]):
        return "cross_border_transfer"
    if any(x in content for x in ["fraud", "scam", "security", "hold", "risk"]):
        return "fraud_assurance"
    if any(x in content for x in ["wallet", "apple pay", "google pay", "checkout", "tap"]):
        return "wallet_or_checkout"
    if any(x in content for x in ["business", "payroll", "merchant"]):
        return "business_payment"
    if any(x in content for x in ["delay", "slow", "instant", "speed", "pending", "transfer"]):
        return "domestic_transfer_speed"
    return "general_payments"


_USE_CASE_LABELS = {
    "cross_border_transfer": "Cross-Border Transfers",
    "fraud_assurance": "Fraud Confidence",
    "wallet_or_checkout": "Wallet And Checkout",
    "business_payment": "Business Payments",
    "domestic_transfer_speed": "Domestic Transfer Speed",
    "general_payments": "General Payments",
}


def _use_case_label(use_case: str) -> str:
    return _USE_CASE_LABELS.get(use_case, use_case.replace("_", " ").title())


def _corroboration_label(unique_domains: int) -> str:
    if unique_domains >= 3:
        return "strong"
    if unique_domains == 2:
        return "moderate"
    return "early"


_CLUSTER_STOPWORDS = {
    "the",
    "and",
    "for",
    "with",
    "that",
    "this",
    "from",
    "into",
    "their",
    "about",
    "your",
    "have",
    "has",
    "had",
    "are",
    "was",
    "were",
    "will",
    "would",
    "should",
    "could",
    "can",
    "but",
    "not",
    "you",
    "they",
    "its",
    "it's",
    "canada",
    "payment",
    "payments",
}


def _normalize_date_value(raw_date: str) -> str:
    value = (raw_date or "").strip()
    if not value:
        return "unknown"
    iso_match = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", value)
    if iso_match:
        return iso_match.group(1)
    return value[:24]


def _extract_date_from_url(url: str) -> str:
    """Layer 1: extract publication date from common URL path patterns.

    Covers most news sites that embed /YYYY/MM/DD/ in their URL structure
    (CBC, Globe and Mail, TechCrunch, Reuters, etc.).
    Returns ISO "YYYY-MM-DD" string or empty string if no date found.
    """
    # /YYYY/MM/DD/ path pattern
    m = re.search(r"/(\d{4})/(\d{1,2})/(\d{1,2})(?:/|$)", url)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    # YYYY-MM-DD or YYYY_MM_DD in URL query string or path segment
    m = re.search(r"(?<!\d)(\d{4})[_\-](\d{2})[_\-](\d{2})(?!\d)", url)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return ""


def _extract_date_from_snippet(text: str) -> str:
    """Layer 2: extract publication date from snippet/body text.

    Handles ISO dates and common English date phrases already present in
    the search snippet (e.g. "Published April 10, 2025" or "2025-03-15").
    Returns ISO "YYYY-MM-DD" string or empty string if nothing matched.
    """
    if not text:
        return ""
    # ISO date in text
    m = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", text)
    if m:
        return m.group(1)
    # "Month DD, YYYY" or "Month DD YYYY"
    _months = (
        "January|February|March|April|May|June|July|August|"
        "September|October|November|December|"
        "Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec"
    )
    m = re.search(
        rf"\b({_months})\s+(\d{{1,2}}),?\s+(\d{{4}})\b", text, re.IGNORECASE
    )
    if m:
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                dt = datetime.strptime(
                    f"{m.group(1)} {m.group(2)} {m.group(3)}", fmt
                )
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    return ""


def _extract_keywords(text: str, max_keywords: int = 6) -> list[str]:
    tokens = re.findall(r"[a-z0-9]+", (text or "").lower())
    keywords: list[str] = []
    for token in tokens:
        if len(token) < 3 or token in _CLUSTER_STOPWORDS:
            continue
        if token not in keywords:
            keywords.append(token)
        if len(keywords) >= max_keywords:
            break
    return keywords


def _cluster_key_from_components(brands: str, use_case: str, keywords: list[str]) -> str:
    keyword_key = ",".join(sorted(keywords[:4])) if keywords else "nokeywords"
    return f"{brands}|{use_case}|{keyword_key}"


def _token_overlap_ratio(tokens_a: set[str], tokens_b: set[str]) -> float:
    if not tokens_a or not tokens_b:
        return 0.0
    overlap = len(tokens_a.intersection(tokens_b))
    return overlap / float(min(len(tokens_a), len(tokens_b)))


def _cluster_mentions(mentions: list[dict]) -> list[dict]:
    clusters: list[dict] = []
    for mention in mentions:
        mention_tokens = set(mention.get("keywords", []))
        chosen_cluster = None

        for cluster in clusters:
            rep = cluster["representative"]
            rep_tokens = set(rep.get("keywords", []))
            overlap = _token_overlap_ratio(mention_tokens, rep_tokens)

            same_brand = mention.get("brands", "Unknown") == rep.get("brands", "Unknown")
            same_use_case = mention.get("use_case", "general_payments") == rep.get("use_case", "general_payments")

            # Anti-fragmentation guard:
            # - strong overlap always merges
            # - fallback merge for same brand + use-case with moderate overlap
            if overlap >= 0.6 or (same_brand and same_use_case and overlap >= 0.35):
                chosen_cluster = cluster
                break

        if chosen_cluster is None:
            chosen_cluster = {
                "cluster_key": mention["cluster_key"],
                "mentions": [],
                "representative": mention,
            }
            clusters.append(chosen_cluster)

        chosen_cluster["mentions"].append(mention)
        if len(mention.get("snippet", "")) > len(chosen_cluster["representative"].get("snippet", "")):
            chosen_cluster["representative"] = mention

    summarized = []
    for idx, cluster in enumerate(clusters, 1):
        cluster_mentions = cluster["mentions"]
        rep = cluster["representative"]
        domains = sorted({urlparse(m.get("url", "")).netloc.replace("www.", "") for m in cluster_mentions if m.get("url")})
        timeframes = sorted({m.get("timeframe", "unknown") for m in cluster_mentions})
        known_dates = sorted(
            {
                d for d in (m.get("date", "unknown") for m in cluster_mentions)
                if d and d != "unknown"
            }
        )
        dated_count = sum(1 for m in cluster_mentions if m.get("date", "unknown") != "unknown")
        if known_dates:
            date_span = known_dates[0] if len(known_dates) == 1 else f"{known_dates[0]} to {known_dates[-1]}"
        else:
            date_span = "unknown"
        summarized.append(
            {
                "story_id": f"S{idx}",
                "archetype_hint": _use_case_label(rep.get("use_case", "general_payments")),
                "brands": rep.get("brands", "Unknown"),
                "article_count": len(cluster_mentions),
                "unique_domains": len(domains),
                "corroboration": _corroboration_label(len(domains)),
                "timeframes_present": ", ".join(timeframes) if timeframes else "unknown",
                "dated_count": dated_count,
                "date_span": date_span,
                "sample_urls": [m.get("url", "") for m in cluster_mentions[:3] if m.get("url")],
                "sample_snippet": rep.get("snippet", "")[:220],
            }
        )

    summarized.sort(key=lambda c: (c["article_count"], c["unique_domains"], c["dated_count"]), reverse=True)
    return summarized


def _extract_platform_context(link: str) -> dict[str, str]:
    """Extract persona-relevant metadata from mention URLs."""
    url = (link or "").lower()
    ctx = {
        "subreddit": "",
        "forum_section": "",
        "platform_demo_hint": "",
    }

    subreddit_match = re.search(r"reddit\.com/r/([a-z0-9_]+)", url)
    if subreddit_match:
        subreddit = subreddit_match.group(1)
        subreddit_hints = {
            "personalfinancecanada": "personal finance consumer, likely 25-45",
            "canadianinvestor": "investor, likely 30-55",
            "canada": "general Canadian public",
            "ontario": "Ontario resident",
        }
        ctx["subreddit"] = subreddit
        ctx["platform_demo_hint"] = subreddit_hints.get(
            subreddit,
            "Reddit community user, likely detail-oriented and price-sensitive",
        )
        return ctx

    if "redflagdeals.com" in url:
        ctx["forum_section"] = "RedFlagDeals"
        ctx["platform_demo_hint"] = "deal-seeking consumer, budget-conscious, likely 25-45"
        return ctx

    if "x.com" in url or "twitter.com" in url:
        ctx["forum_section"] = "X/Twitter"
        ctx["platform_demo_hint"] = "social media user, skews 20-40, more reactive"
        return ctx

    if "forum" in url or "community" in url:
        ctx["forum_section"] = "Forum/Community"
        ctx["platform_demo_hint"] = "community forum user, likely troubleshooting-focused"

    return ctx


def _tbs_to_timelimit(tbs: str) -> str | None:
    mapping = {"qdr:d": "d", "qdr:w": "w", "qdr:m": "m", "qdr:y": "y"}
    return mapping.get(tbs) if tbs else None


def _resolve_relative_date(date_str: str, *, tbs: str = "") -> str:
    """Normalize DDG date strings to "Month DD, YYYY" format (UTC dates).

    DDG text() returns relative strings like '3 weeks ago', '2 months ago'.
    DDG news() returns ISO timestamps or pre-formatted strings.
    All dates are kept in UTC — the native timezone of Reddit's created_utc
    and DDG's timestamps — so the displayed date matches when the content
    was actually published.
    Returns empty string if no date is available.
    """
    if not date_str:
        return date_str
    # ISO dates: "2025-04-10", "2025-04-10T14:22:00+00:00", "2025-04-10 14:22:00"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", date_str)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)), tzinfo=timezone.utc)
            return dt.strftime("%B %d, %Y")
        except ValueError:
            pass
    # RFC 2822: "Mon, 14 Apr 2026 01:30:00 +0000"
    try:
        from email.utils import parsedate_to_datetime as _pdt
        dt = _pdt(date_str).astimezone(timezone.utc)
        return dt.strftime("%B %d, %Y")
    except Exception:
        pass
    # Named-month: "April 14, 2026" / "Apr 14, 2026" / "14 Apr 2026"
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%B %d, %Y")
        except ValueError:
            pass
    # Parse relative strings using UTC now so the computed date matches the source
    now = datetime.now(timezone.utc)
    m = re.match(r"(\d+)\s+(day|week|month|year)s?\s+ago", date_str.strip(), re.IGNORECASE)
    if m:
        n = int(m.group(1))
        unit = m.group(2).lower()
        if unit == "day":
            dt = now - timedelta(days=n)
        elif unit == "week":
            dt = now - timedelta(weeks=n)
        elif unit == "month":
            dt = now - timedelta(days=n * 30)
        else:  # year
            dt = now - timedelta(days=n * 365)
        return dt.strftime("%B %d, %Y")
    if re.match(r"\d+\s+(hour|minute)s?\s+ago", date_str.strip(), re.IGNORECASE):
        return now.strftime("%B %d, %Y")
    return date_str  # unknown format — keep as-is


async def _fetch_meta_date(url: str, client: httpx.AsyncClient) -> str:
    """Layer 3: scrape HTML <head> for Open Graph / JSON-LD / <time> publication date.

    Returns "Month DD, YYYY" string or empty string on failure.
    Only reads the first 25 KB of the response — enough to cover most <head> blocks.
    """
    try:
        resp = await client.get(
            url,
            follow_redirects=True,
            timeout=4.0,
            headers={"User-Agent": "Mozilla/5.0 (compatible; InteracIntelBot/1.0)"},
        )
        if resp.status_code != 200:
            return ""
        head_html = resp.text[:25000]
    except Exception:
        return ""

    patterns = [
        # Open Graph article:published_time (both attribute orderings)
        r'<meta[^>]+property=["\']article:published_time["\'][^>]+content=["\']([\d\-T:+Z]+)["\']',
        r'<meta[^>]+content=["\']([\d\-T:+Z]+)["\'][^>]+property=["\']article:published_time["\']',
        # Generic pubdate / DC.date meta tags
        r'<meta[^>]+name=["\'](?:pubdate|DC\.date)["\'][^>]+content=["\']([\d\-T:+Z]+)["\']',
        # Common modern meta date keys
        r'<meta[^>]+name=["\'](?:date|publish-date|article\.published|parsely-pub-date|sailthru\.date)["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+property=["\'](?:article:modified_time|og:updated_time)["\'][^>]+content=["\']([^"\']+)["\']',
        # JSON-LD datePublished
        r'"datePublished"\s*:\s*"([\d\-T:+Z]+)"',
        # HTML5 <time datetime="...">
        r'<time[^>]+datetime=["\']([^"\']+)["\']',
    ]
    for pat in patterns:
        m = re.search(pat, head_html, re.IGNORECASE)
        if m:
            resolved = _resolve_relative_date(m.group(1))
            if resolved and not re.match(r"^\d{4}-\d{2}-\d{2}T", resolved):
                # _resolve_relative_date already formatted it as "Month DD, YYYY"
                return resolved
            # ISO timestamp that _resolve_relative_date parsed cleanly
            if resolved:
                return resolved
    return ""


async def _enrich_dates_from_meta(
    mentions: list[dict], max_fetches: int = 20
) -> None:
    """Layer 3 batch enrichment: fetch HTML meta dates for undated mentions in-place.

    Skips mentions that already have a date. Caps HTTP requests at max_fetches
    to bound latency (each request has a 4 s timeout, all run concurrently).
    """
    undated = [
        m for m in mentions
        if not m.get("date") and m.get("link")
    ]
    targets = undated[:max_fetches]
    if not targets:
        return

    async with httpx.AsyncClient() as client:
        results = await asyncio.gather(
            *[_fetch_meta_date(m["link"], client) for m in targets],
            return_exceptions=True,
        )

    gained = 0
    for mention, result in zip(targets, results):
        if isinstance(result, str) and result:
            mention["date"] = result
            gained += 1

    logger.info(f"[date-enrich-L3] fetched {len(targets)} undated URLs, gained {gained} dates")


def _canonical_url_for_date_lookup(url: str) -> str:
    """Normalize URL for matching mention links to email bullet Source URLs."""
    u = (url or "").strip().rstrip(".,)")
    if not u:
        return ""
    try:
        p = urlparse(u)
    except Exception:
        return ""
    host = (p.netloc or "").lower()
    path = (p.path or "").rstrip("/")
    query = p.query or ""
    if "reddit.com" in host:
        host = "reddit.com"
        return urlunparse(("https", host, path, "", "", ""))
    if host.startswith("www."):
        host = host[4:]
    return urlunparse(("https", host, path, "", query, ""))


def _build_url_date_map_from_mentions(*buckets: list[dict]) -> dict[str, str]:
    out: dict[str, str] = {}
    for bucket in buckets:
        for m in bucket:
            link = (m.get("link") or "").strip()
            date = (m.get("date") or "").strip()
            if not link or not date:
                continue
            key = _canonical_url_for_date_lookup(link)
            if key:
                out.setdefault(key, date)
    return out


def _parse_display_date_utc(date_str: str) -> datetime | None:
    """Parse formatted mention date like 'April 14, 2026' into UTC datetime."""
    value = (date_str or "").strip()
    if not value or value.lower() == "unknown":
        return None
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def _filter_recent_dated_mentions(
    mentions: list[dict], *, max_age_days: int
) -> list[dict]:
    """Keep only mentions with a known date newer than max_age_days."""
    if max_age_days <= 0:
        return mentions
    cutoff = datetime.now(timezone.utc) - timedelta(days=max_age_days)
    kept: list[dict] = []
    dropped_undated = 0
    dropped_stale = 0
    for m in mentions:
        d = _parse_display_date_utc(m.get("date", ""))
        if d is None:
            dropped_undated += 1
            continue
        if d < cutoff:
            dropped_stale += 1
            continue
        kept.append(m)
    if dropped_undated or dropped_stale:
        logger.info(
            f"[recency-filter] kept={len(kept)} dropped_undated={dropped_undated} "
            f"dropped_stale={dropped_stale} max_age_days={max_age_days}"
        )
    return kept


_ETRANSFER_SIGNAL_TOKENS = [
    "e-transfer", "etransfer", "interac", "auto-deposit", "autodeposit",
    "fraud", "scam", "hold", "pending", "declined", "limit", "delay", "reversal",
]
_CANADIAN_BANK_NAMES = [
    "td bank", "rbc", "scotiabank", "bmo", "cibc", "national bank",
    "tangerine", "eq bank", "simplii", "desjardins", "hsbc canada",
]
_DOLLAR_AMOUNT_RE = re.compile(r"\$\s*\d[\d,]*")

# Insight-signal words — indicate a real observation, comparison, or experience
# rather than a generic statement. These boost the quality score.
_INSIGHT_SIGNAL_TOKENS = [
    "switched", "switching", "instead of", "better than", "worse than",
    "compared to", "prefer", "moved to", "stopped using", "started using",
    "annoying", "frustrated", "surprised", "realized", "just happened",
    "required", "prompted", "blocked me", "rejected", "wouldn't let",
    "finally", "used to", "no longer", "can't believe", "disappointed",
    "love that", "hate that", "wish", "should", "needs to",
]

# Low-insight patterns — posts where e-transfer / payment is mentioned only
# in passing with no real observation or experience attached.
_LOW_INSIGHT_RE = re.compile(
    r"(equivalent (to|of) (interac|e-transfer)|"
    r"they (don't|dont|do not) (have|use) (interac|e-transfer)|"
    r"(interac|e-transfer) (or|and) (cash|cheque|check|credit card)|"
    r"payable.{0,30}(interac|e-transfer)|"  # "payable via e-transfer" — incidental
    r"(prize|reward|payment).{0,40}(interac|e-transfer))",  # contest/prize mentions
    re.IGNORECASE,
)

_COMPETITOR_BRANDS = [
    "wise", "paypal", "wealthsimple", "koho", "apple pay", "google pay",
    "revolut", "neo financial", "venmo", "zelle", "cash app", "square", "stripe",
]
_COMPETITOR_EVENT_TOKENS = [
    "launch", "launched", "rollout", "rolled out", "new feature", "pricing",
    "fee", "fees", "partnership", "expands", "expansion", "introduces", "update",
]
_LOW_SIGNAL_PATTERNS = [
    "what do you think", "is this legit", "anyone else", "help please", "i'm new",
]


def _mention_quality_score(mention: dict, section: str) -> float:
    """Score mention quality using platform-specific signals, normalized to a common scale.

    Reddit:  upvotes (capped 500) + comments (capped 100) → 0–4 pts
    Twitter: (likes + retweets×2 + replies) (capped 1000) → 0–4 pts
    Both:    snippet length, keyword signals, dollar amounts, bank names
    """
    text = " ".join(
        [
            mention.get("title", ""),
            mention.get("snippet", ""),
        ]
    ).lower()
    link = (mention.get("link", "") or "").lower()
    score = 0.0

    channel, _ = _classify_channel_and_source(link)
    source_tier = _source_quality_tier(link, channel)
    if source_tier == "tier1_user_generated":
        score += 1.5
    elif source_tier == "tier2_reported":
        score += 1.0

    # ── Platform-specific engagement signals (normalized to 0–4 pts each) ──
    if "reddit.com" in link:
        # Upvotes: community validation. Cap at 500 (most viral posts).
        upvotes = min(float(mention.get("score", 0)), 500.0)
        score += (upvotes / 500.0) * 3.0
        # Comments: depth of discussion. Cap at 100.
        comments = min(float(mention.get("num_comments", 0)), 100.0)
        score += (comments / 100.0) * 1.0

    elif mention.get("source") == "X/Twitter":
        # Likes + retweets (2× weight — retweet = active amplification) + replies.
        # Cap at 1000 total weighted engagement.
        likes = mention.get("_likes") or mention.get("_engagement") or 0
        retweets = mention.get("_retweets") or 0
        replies = mention.get("_replies") or 0
        weighted = min(float(likes + retweets * 2 + replies), 1000.0)
        score += (weighted / 1000.0) * 3.0
        # Views: weak signal (passive), only helps if very high (>10k)
        views = mention.get("_views") or 0
        if views >= 10000:
            score += 0.5
        elif views >= 1000:
            score += 0.2

    # ── Content quality signals (platform-neutral) ──
    snippet_len = len((mention.get("snippet", "") or "").strip())
    if snippet_len >= 90:
        score += 1.0
    if snippet_len >= 200:
        score += 0.5

    # Dollar amounts signal a real personal experience (e.g. "my $2,400 transfer")
    if _DOLLAR_AMOUNT_RE.search(text):
        score += 1.0

    # Canadian bank names signal relevant context
    if any(bank in text for bank in _CANADIAN_BANK_NAMES):
        score += 0.5

    if any(p in text for p in _LOW_SIGNAL_PATTERNS):
        score -= 1.5
    if "?" in text and snippet_len < 120:
        score -= 0.5

    if section == "etransfer":
        if any(tok in text for tok in _ETRANSFER_SIGNAL_TOKENS):
            score += 1.5
        else:
            score -= 3.0
        # Insight quality: reward posts with real user experience signals
        if any(tok in text for tok in _INSIGHT_SIGNAL_TOKENS):
            score += 1.0
        # Penalise low-insight patterns (passing references, prize mentions, etc.)
        if _LOW_INSIGHT_RE.search(text):
            score -= 2.0
    else:  # competitor
        has_brand = any(tok in text for tok in _COMPETITOR_BRANDS)
        if has_brand:
            score += 1.0
        else:
            score -= 2.0
        if channel == "press" and not any(tok in text for tok in _COMPETITOR_EVENT_TOKENS):
            score -= 1.5
        if any(tok in text for tok in _COMPETITOR_EVENT_TOKENS):
            score += 1.0
        # Insight quality: reward comparison/switch posts in competitor section too
        if any(tok in text for tok in _INSIGHT_SIGNAL_TOKENS):
            score += 0.5

    return score


def _quality_gate_mentions(
    mentions: list[dict], *, section: str, threshold: float
) -> list[dict]:
    """Keep only quality mentions; preserve ranking by score."""
    if not QUALITY_STRICT:
        return mentions
    kept: list[dict] = []
    dropped = 0
    for m in mentions:
        q = _mention_quality_score(m, section)
        if q >= threshold:
            kept.append(m)
        else:
            dropped += 1
    if dropped:
        logger.info(
            f"[quality-filter] section={section} kept={len(kept)} dropped={dropped} threshold={threshold}"
        )
    return kept


async def web_search(
    query: str,
    search_type: str = "search",
    max_results: int = 5,
    tbs: str = "qdr:w",
) -> list[dict]:
    timelimit = _tbs_to_timelimit(tbs)

    def _run_search() -> list[dict]:
        with DDGS() as ddgs:
            if search_type == "news":
                raw = list(ddgs.news(query, max_results=max_results, timelimit=timelimit))
            else:
                raw = list(ddgs.text(query, max_results=max_results, timelimit=timelimit))
        normalized = []
        for item in raw:
            link = item.get("href", "") or item.get("url", "")
            # text() uses "published" ("2 weeks ago"); news() uses "date" (ISO).
            raw_date = item.get("date", "") or item.get("published", "")
            # Layer 1: try URL path pattern when API field is empty
            if not raw_date:
                raw_date = _extract_date_from_url(link)
            # Layer 2: try snippet text when URL pattern also fails
            if not raw_date:
                raw_date = _extract_date_from_snippet(
                    item.get("body", "") or item.get("snippet", "")
                )
            normalized.append({
                "title": item.get("title", ""),
                "snippet": item.get("body", "") or item.get("snippet", ""),
                "link": link,
                "source": item.get("source", search_type),
                # _resolve_relative_date converts any date string to "Month DD, YYYY".
                # Pass tbs so empty dates get an approximate fallback from the search window.
                "date": _resolve_relative_date(raw_date, tbs=tbs),
            })
        return normalized

    try:
        results = await asyncio.to_thread(_run_search)
    except Exception as e:
        err = f"DDG exception for [{search_type}] '{query[:40]}': {type(e).__name__}: {e}"
        logger.error(err)
        _search_errors.append(err)
        return []

    logger.info(f"DDG [{search_type}] '{query}' tbs={tbs!r} -> {len(results)} results")
    return results


async def _search_twitter_io(query: str, max_results: int = 10) -> list[dict]:
    """Search X/Twitter via twitterapi.io. Returns [] if key not set or on error."""
    if not TWITTERAPI_IO_KEY:
        return []
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                TWITTERAPI_IO_URL,
                params={"query": query, "queryType": "Latest"},
                headers={"x-api-key": TWITTERAPI_IO_KEY},
            )
            resp.raise_for_status()
            data = resp.json()
    except Exception as e:
        logger.warning(f"[twitterapi.io] request failed for '{query}': {e}")
        return []

    _PROMO_RE = re.compile(
        r"\b(sign up|download|click here|visit us|check out|get started|"
        r"use code|promo|discount|affiliate|refer a friend|sign-up bonus|"
        r"join us|follow us|learn more|limited time|giveaway|contest|"
        r"prize|winner|enter to win)\b",
        re.IGNORECASE,
    )
    _CRYPTO_JUNK_RE = re.compile(
        r"\b(stablecoin|on-chain|off-ramp|web3|nft|defi|staking|"
        r"\$[A-Z]{3,6}\b|@daimo|@cadcstablecoin|loonfinance|crypto bridge|"
        r"smart contract|blockchain)\b",
        re.IGNORECASE,
    )
    _PRONOUN_RE = re.compile(
        r"\b(i|my|me|we|our|i've|i'm|i'd|i'll|we've|we're)\b", re.IGNORECASE
    )
    _MENTION_RE = re.compile(r"@\w+")

    results = []
    for tweet in (data.get("tweets") or []):
        if len(results) >= max_results:
            break
        text = (tweet.get("text") or "").strip()
        url = (tweet.get("url") or "").strip()
        created_at = (tweet.get("createdAt") or "").strip()
        author = tweet.get("author") or {}
        username = author.get("userName") or author.get("name") or ""

        if not text or not url:
            continue
        # Skip short tweets — no room for insight
        if len(text) < 100:
            continue
        # Skip promotional content
        if _PROMO_RE.search(text):
            continue
        # Skip crypto/web3 tangents — they mention e-Transfer but aren't about user experience
        if _CRYPTO_JUNK_RE.search(text):
            continue
        # Skip tweets with too many @mentions (bot/promotional/ecosystem posts)
        if len(_MENTION_RE.findall(text)) > 2:
            continue

        likes = tweet.get("likeCount") or 0
        retweets = tweet.get("retweetCount") or 0

        # Balanced bar: engagement ≥ 1 AND at least one insight signal,
        # OR engagement = 0 BUT at least two strong signals (for fresh tweets
        # that haven't been liked yet). Value filter handles the rest.
        has_engagement = (likes + retweets) >= 1
        has_pronoun = bool(_PRONOUN_RE.search(text))
        has_dollar = bool(_DOLLAR_AMOUNT_RE.search(text))
        has_bank = any(bank in text.lower() for bank in _CANADIAN_BANK_NAMES)
        has_question = "?" in text
        has_insight_token = any(tok in text.lower() for tok in _INSIGHT_SIGNAL_TOKENS)

        concrete_signals = sum([has_pronoun, has_dollar, has_bank, has_question, has_insight_token])

        if has_engagement:
            if concrete_signals < 1:
                continue
        else:
            # No engagement — require stronger content signal
            if concrete_signals < 2:
                continue

        results.append({
            "title": f"Tweet by @{username}" if username else "Tweet",
            "snippet": text,
            "link": url,
            "date": _resolve_relative_date(created_at) if created_at else "",
            "source": "X/Twitter",
            "_likes": likes,
            "_retweets": retweets,
            "_replies": tweet.get("replyCount") or 0,
            "_views": tweet.get("viewCount") or 0,
            "_engagement": likes + retweets,  # kept for backwards compat
        })
    return results


async def search_twitter(query: str, max_results: int = 5, tbs: str = "qdr:w") -> list[dict]:
    """Search X/Twitter. Uses twitterapi.io if key is set, otherwise falls back to DDG."""
    # Prefer real Twitter API when available
    if TWITTERAPI_IO_KEY:
        return await _search_twitter_io(query, max_results=max_results)
    # DDG fallback — skips if query already has a site: restriction
    if _has_site_restriction(query):
        return []
    base_results = await web_search(
        f"{query} site:x.com OR site:twitter.com",
        "search",
        max_results=max_results,
        tbs=tbs,
    )
    for r in base_results:
        r["source"] = "X/Twitter"
    return base_results


async def search_google_news(query: str, max_results: int = 10, days_back: int = 30) -> list[dict]:
    """Fetch Google News RSS (CA-geo). Returns normalized mention dicts with reliable dates.
    days_back: inject after:YYYY-MM-DD into query to limit to recent articles."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days_back)).strftime("%Y-%m-%d")
    dated_query = f"{query} after:{cutoff}"
    url = (
        f"https://news.google.com/rss/search?q={quote(dated_query)}"
        f"&hl=en-CA&gl=CA&ceid=CA:en"
    )
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(url, follow_redirects=True)
            resp.raise_for_status()
    except Exception as e:
        logger.warning(f"[google-news] failed for '{query}': {e}")
        return []

    items = []
    try:
        root = ET.fromstring(resp.text)
        for item in root.findall(".//item")[:max_results]:
            link = item.findtext("link", "") or ""
            if not link:
                continue
            title   = item.findtext("title", "") or ""
            pub_raw = item.findtext("pubDate", "") or ""
            desc    = item.findtext("description", "") or ""
            source_el = item.find("source")
            source_name = (source_el.text if source_el is not None else "Google News") or "Google News"

            date_str = ""
            if pub_raw:
                try:
                    dt = parsedate_to_datetime(pub_raw)
                    date_str = dt.strftime("%B %d, %Y")
                except Exception:
                    pass

            items.append({
                "title":   title,
                "link":    link,
                "snippet": re.sub(r"<[^>]+>", "", desc).strip(),
                "date":    date_str,
                "source":  source_name,
                "_fetch_method": "google_news_rss",
            })
    except ET.ParseError as e:
        logger.warning(f"[google-news] parse error for '{query}': {e}")
        return []

    return items


_REDDIT_HEADERS = {
    "User-Agent": "python:interac.intelligence.bot:v1.0 (by /u/interac_intel_bot)"
}

_REDDIT_THREAD_PATH = re.compile(r"^/r/[^/]+/comments/[a-z0-9]+", re.I)


def _is_reddit_thread_url(url: str) -> bool:
    try:
        p = urlparse(url)
    except Exception:
        return False
    if "reddit.com" not in (p.netloc or "").lower():
        return False
    return bool(_REDDIT_THREAD_PATH.match(p.path or ""))


async def _reddit_json_created_str(url: str, client: httpx.AsyncClient) -> str:
    """Fetch thread created_utc via Reddit's .json suffix. Returns '' on failure."""
    base = url.split("#")[0].split("?")[0].rstrip("/")
    jurl = base if base.lower().endswith(".json") else f"{base}.json"
    try:
        r = await client.get(jurl, headers=_REDDIT_HEADERS, timeout=12.0)
        if r.status_code != 200:
            return ""
        data = r.json()
    except Exception:
        return ""
    created = None
    if isinstance(data, list) and data:
        children = data[0].get("data", {}).get("children", [])
        if children:
            created = children[0].get("data", {}).get("created_utc")
    if created is None:
        return ""
    try:
        post_dt = datetime.fromtimestamp(float(created), tz=timezone.utc)
        return post_dt.strftime("%B %d, %Y")
    except Exception:
        return ""


async def _enrich_dates_from_reddit_json(
    mentions: list[dict], *, max_fetches: int = 40
) -> None:
    """Fill missing mention dates for Reddit thread URLs using the public .json endpoint."""
    targets: list[tuple[str, str]] = []
    seen_keys: set[str] = set()
    for m in mentions:
        link = (m.get("link") or "").strip()
        if (m.get("date") or "").strip():
            continue
        if not link or not _is_reddit_thread_url(link):
            continue
        key = _canonical_url_for_date_lookup(link)
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        targets.append((key, link))
        if len(targets) >= max_fetches:
            break

    if not targets:
        return

    async with httpx.AsyncClient(follow_redirects=True) as client:
        results = await asyncio.gather(
            *[_reddit_json_created_str(link, client) for _key, link in targets],
            return_exceptions=True,
        )

    key_to_date: dict[str, str] = {}
    for (key, _link), res in zip(targets, results):
        if isinstance(res, str) and res:
            key_to_date[key] = res

    if not key_to_date:
        logger.info(f"[date-enrich-reddit] fetched {len(targets)} undated Reddit URLs, gained 0 dates")
        return

    for m in mentions:
        if (m.get("date") or "").strip():
            continue
        lk = (m.get("link") or "").strip()
        if not lk:
            continue
        d = key_to_date.get(_canonical_url_for_date_lookup(lk))
        if d:
            m["date"] = d

    logger.info(
        f"[date-enrich-reddit] fetched {len(targets)} undated Reddit URLs, "
        f"gained {len(key_to_date)} dates"
    )


def _parse_reddit_post(post: dict, cutoff_ts: float) -> dict | None:
    """Parse a Reddit post dict into our standard format. Returns None if filtered out."""
    created = post.get("created_utc", 0)
    if not created or created < cutoff_ts:
        return None
    title = (post.get("title", "") or "").strip()
    selftext = (post.get("selftext", "") or "").strip()
    if selftext in ("[deleted]", "[removed]"):
        selftext = ""
    permalink = post.get("permalink", "")
    subreddit_name = post.get("subreddit_display_name", post.get("subreddit", ""))
    post_dt = datetime.fromtimestamp(created, tz=timezone.utc)
    return {
        "title": title,
        "snippet": selftext[:900] if selftext else title,
        "link": f"https://www.reddit.com{permalink}",
        "source": f"Reddit/r/{subreddit_name}" if subreddit_name else "Reddit",
        "date": post_dt.strftime("%B %d, %Y"),
        "score": post.get("score", 0),
        "num_comments": post.get("num_comments", 0),
        "permalink": permalink,
    }


async def _reddit_get(url: str, params: dict) -> dict | None:
    """GET a Reddit JSON endpoint using httpx (urllib gets 403'd by Reddit's TLS check)."""
    try:
        async with httpx.AsyncClient(timeout=15, headers=_REDDIT_HEADERS) as client:
            r = await client.get(url, params=params)
        if r.status_code != 200:
            logger.warning(f"Reddit API {r.status_code} for {url}")
            return None
        return r.json()
    except Exception as e:
        logger.warning(f"Reddit API failed for {url}: {type(e).__name__}: {e}")
        return None


async def fetch_reddit_comments(permalink: str, max_comments: int = 6) -> list[str]:
    """Fetch top comments from a post by appending .json to its permalink URL.

    This is the 'add .json to any Reddit URL' trick — gives full post + comments.
    Only returns comments with score >= 2.
    """
    url = f"https://www.reddit.com{permalink}.json"
    try:
        async with httpx.AsyncClient(timeout=12, headers=_REDDIT_HEADERS) as client:
            r = await client.get(url, params={"limit": max_comments, "sort": "top"})
        if r.status_code != 200:
            return []
        data = r.json()
        # data[0] = post listing, data[1] = comment listing
        comments = data[1]["data"]["children"]
        result = []
        for c in comments[:max_comments]:
            d = c.get("data", {})
            body = (d.get("body", "") or "").strip()
            cscore = d.get("score", 0)
            if body and body not in ("[deleted]", "[removed]") and cscore >= 2:
                result.append(body[:350])
        return result
    except Exception as e:
        logger.debug(f"Comment fetch failed for {permalink}: {e}")
        return []


async def search_reddit_posts(
    query: str,
    subreddit: str = "",
    max_results: int = 15,
    days_back: int = 60,
    min_score: int = 1,
    enrich_comments: bool = False,
) -> list[dict]:
    """Search Reddit posts via the public JSON API using httpx.

    - Exact UTC timestamps → real dates
    - Full post selftext → better quote material
    - Optional comment enrichment: for posts with score >= 5, fetches top comment
      (the .json trick) and appends it to the snippet so the LLM has community
      reactions, not just the OP's question
    """
    if subreddit:
        url = f"https://www.reddit.com/r/{subreddit}/search.json"
        params: dict = {
            "q": query, "sort": "new", "t": "year",
            "limit": min(max_results * 2, 25), "restrict_sr": "1", "type": "link",
        }
    else:
        url = "https://www.reddit.com/search.json"
        params = {
            "q": query, "sort": "new", "t": "year",
            "limit": min(max_results * 2, 25), "type": "link",
        }

    cutoff_ts = (datetime.now(timezone.utc) - timedelta(days=days_back)).timestamp()
    data = await _reddit_get(url, params)
    if not data:
        return []

    posts = []
    for child in data.get("data", {}).get("children", []):
        parsed = _parse_reddit_post(child.get("data", {}), cutoff_ts)
        if parsed and parsed["score"] >= min_score:
            posts.append(parsed)

    posts.sort(key=lambda x: x["score"], reverse=True)
    posts = posts[:max_results]

    # Enrich top posts with their best comment via the .json permalink trick
    if enrich_comments:
        enrichment_tasks = []
        for p in posts:
            if p["score"] >= 5 and p.get("permalink"):
                enrichment_tasks.append((p, fetch_reddit_comments(p["permalink"], max_comments=4)))
            else:
                enrichment_tasks.append((p, None))

        for post, coro in enrichment_tasks:
            if coro is None:
                continue
            comments = await coro
            if comments:
                # Prepend the best comment to give the LLM community reaction context
                post["snippet"] = post["snippet"] + "\n\nTop community reply: " + comments[0]

    return posts


async def browse_subreddit_new(
    subreddit: str,
    keyword: str,
    days_back: int = 30,
    limit: int = 100,
) -> list[dict]:
    """Browse a subreddit's /new feed and filter by keyword — no search API bias.

    Useful when you want all recent posts mentioning a term, not just what
    Reddit's search algorithm returns.
    """
    url = f"https://www.reddit.com/r/{subreddit}/new.json"
    cutoff_ts = (datetime.now(timezone.utc) - timedelta(days=days_back)).timestamp()
    data = await _reddit_get(url, {"limit": limit})
    if not data:
        return []

    kw = keyword.lower()
    posts = []
    for child in data.get("data", {}).get("children", []):
        post_data = child.get("data", {})
        title = (post_data.get("title", "") or "").lower()
        selftext = (post_data.get("selftext", "") or "").lower()
        if kw not in title and kw not in selftext:
            continue
        parsed = _parse_reddit_post(post_data, cutoff_ts)
        if parsed:
            posts.append(parsed)

    posts.sort(key=lambda x: x["score"], reverse=True)
    return posts


async def _search_diagnostic() -> str:
    """Single test query to diagnose DDG search health."""
    try:
        results = await web_search("Interac e-Transfer", "search", max_results=3, tbs="")
    except Exception as e:
        return f"FAIL: {type(e).__name__}: {e}"
    if not results:
        return "FAIL: no search results returned"
    return f"OK: {len(results)} results, first='{results[0].get('title', 'n/a')[:80]}'"


async def fetch_biweekly_mentions(*, quarterly: bool = False) -> tuple[str, list[dict]]:
    """Fetch mentions for the universal biweekly scan (or a wider quarterly pool).

    Strategy:
    - Reddit API (primary): exact dates, full post content, quality score filter.
      Used for both e-Transfer pain posts and competitor community reactions.
    - DDG (supplement): RFD, X/Twitter, news for things Reddit API can't cover.
    Results split into labelled sections so the LLM assigns them correctly.

    When ``quarterly=True``:
    - Rolling ~90-day window (no biweekly ``sent_urls`` dedupe — quarterly needs full history).
    - Broader DDG time filter and larger per-section caps for the long-form report.
    """
    global last_biweekly_url_dates, last_quarterly_url_dates
    config = load_prompts()
    etransfer_ddg_queries = config.get("etransfer_queries", config.get("biweekly_queries", []))
    competitor_ddg_queries = config.get("competitor_queries", [])

    ddg_tbs = "" if quarterly else "qdr:m"
    max_age_days = 90 if quarterly else MAX_MENTION_AGE_DAYS

    def _rd_days(d: int) -> int:
        return max(d, 90) if quarterly else d

    # Biweekly: skip URLs already surfaced so scans stay fresh. Quarterly: do not
    # dedupe against that list — the report must reflect the full ~90-day window.
    _memory = _load_biweekly_memory()
    previously_sent: set[str] = set()
    if not quarterly:
        previously_sent = set(_memory.get("sent_urls", []))
        if previously_sent:
            logger.info(f"[dedup] filtering against {len(previously_sent)} previously-sent URLs")

    seen_links: set[str] = set(previously_sent)  # pre-seed (empty for quarterly)
    etransfer_social: list[dict] = []
    etransfer_press: list[dict] = []
    competitor_mentions: list[dict] = []

    sem = asyncio.Semaphore(3)  # Max 3 concurrent Reddit API calls

    async def _search(q: str, sub: str, days: int, score: int = 1, comments: bool = False) -> list[dict]:
        async with sem:
            return await search_reddit_posts(
                q, subreddit=sub, max_results=12, days_back=days,
                min_score=score, enrich_comments=comments,
            )

    async def _browse(sub: str, kw: str, days: int) -> list[dict]:
        async with sem:
            return await browse_subreddit_new(sub, kw, days_back=days, limit=100)

    # ── 1. Reddit — e-Transfer community ──
    # Search API for targeted queries + browse /new feed for anything mentioning
    # e-transfer that search might miss (no search-ranking bias on the browse).
    reddit_et_searches = [
        ("e-transfer", "personalfinancecanada"),
        ("interac e-transfer", "personalfinancecanada"),
        ("e-transfer fraud OR scam", ""),
        ("e-transfer problem OR issue OR delay OR complaint", ""),
        ("e-transfer limit OR hold OR declined OR pending", ""),
        ("interac e-transfer", "canada"),
        ("e-transfer", "banking"),
        ("e-transfer", "ontario"),
        ("e-transfer", "toronto"),
        ("interac e-transfer fraud OR scam", "Scams"),
        ("interac e-transfer", "legaladvicecanada"),
        ("e-transfer", "frugalcanada"),
        ("e-transfer help OR problem", "CanadianInvestor"),
    ]
    et_browse = [
        ("personalfinancecanada", "e-transfer", 45),
        ("personalfinancecanada", "interac", 45),
        ("canada", "e-transfer", 30),
        ("ontario", "e-transfer", 30),
        ("Scams", "e-transfer", 45),
        ("frugalcanada", "e-transfer", 60),
    ]

    et_search_tasks = [_search(q, sub, _rd_days(60)) for q, sub in reddit_et_searches]
    et_browse_tasks = [_browse(sub, kw, _rd_days(days)) for sub, kw, days in et_browse]
    et_batches = await asyncio.gather(*et_search_tasks, *et_browse_tasks, return_exceptions=True)

    for batch in et_batches:
        if isinstance(batch, Exception) or not batch:
            continue
        for r in batch:
            link = r.get("link", "")
            if link and link not in seen_links:
                seen_links.add(link)
                r["_fetch_method"] = "reddit_json"
                etransfer_social.append(r)

    logger.info(f"[reddit-api] et-Transfer community returned {len(etransfer_social)} posts")

    # ── 2. Reddit — competitor community reactions (with comment enrichment) ──
    # enrich_comments=True: for posts score>=5, fetches top comment via .json trick
    # so the LLM gets community reaction, not just the OP's question.
    reddit_comp_searches = [
        ("wise money transfer canada", "personalfinancecanada"),
        ("paypal canada send money", "personalfinancecanada"),
        ("wealthsimple cash transfer", "personalfinancecanada"),
        ("koho card banking", "personalfinancecanada"),
        ("best way send money canada interac alternative", "personalfinancecanada"),
        ("apple pay google pay canada", "personalfinancecanada"),
        ("revolut canada", "personalfinancecanada"),
        ("neo financial banking canada", "personalfinancecanada"),
        ("venmo paypal zelle canada", ""),
        ("digital wallet canada", "personalfinancecanada"),
        ("switched from e-transfer", "personalfinancecanada"),
        ("e-transfer alternative", "personalfinancecanada"),
        ("wise vs interac OR paypal vs interac", "personalfinancecanada"),
        ("wise money transfer canada", "canada"),
        ("revolut canada", "canada"),
        ("koho OR wealthsimple cash", "canada"),
    ]
    comp_browse = [
        ("personalfinancecanada", "wise", 90),
        ("personalfinancecanada", "paypal", 90),
        ("personalfinancecanada", "wealthsimple", 60),
        ("personalfinancecanada", "revolut", 90),
        ("canada", "wise", 60),
        ("canada", "koho", 60),
    ]

    comp_search_tasks = [_search(q, sub, _rd_days(180), score=1, comments=True) for q, sub in reddit_comp_searches]
    comp_browse_tasks = [_browse(sub, kw, _rd_days(days)) for sub, kw, days in comp_browse]
    comp_batches = await asyncio.gather(*comp_search_tasks, *comp_browse_tasks, return_exceptions=True)

    for batch in comp_batches:
        if isinstance(batch, Exception) or not batch:
            continue
        for r in batch:
            link = r.get("link", "")
            if link and link not in seen_links:
                seen_links.add(link)
                r["_fetch_method"] = "reddit_json"
                competitor_mentions.append(r)

    # ── 3a. Dedicated X/Twitter e-Transfer searches ──
    # Run these independently so they aren't limited by the DDG query list.
    et_twitter_queries = [
        "interac e-transfer",
        "e-transfer scam OR fraud",
        "e-transfer hold OR declined OR pending",
        "e-transfer complaint OR issue OR broken",
    ]
    for tq in et_twitter_queries:
        # Cap at 4 per query (down from 8) — prevents Twitter flooding the social pool
        x_results = await search_twitter(tq, 4, tbs=ddg_tbs)
        for r in x_results:
            link = r.get("link", "")
            if not link or link in seen_links or _is_blocked_domain(link):
                continue
            seen_links.add(link)
            r["channel"] = "people"
            r["source"] = "X/Twitter"
            r["_fetch_method"] = "ddg_text"
            etransfer_social.append(r)

    # ── 3b. Mandatory Reddit DDG fallback (runs regardless of skip_ddg) ─────────
    # Reddit API is frequently blocked on Railway. These site:reddit.com queries
    # via DDG always run to ensure Reddit content reaches the pool.
    reddit_ddg_fallback = [
        "interac e-transfer site:reddit.com/r/personalfinancecanada",
        "e-transfer problem OR scam site:reddit.com",
        "e-transfer held pending stuck site:reddit.com",
        "bank e-transfer issue site:reddit.com/r/canada",
        "interac auto-deposit site:reddit.com",
        "e-transfer declined OR failed OR error site:reddit.com",
        "TD OR RBC OR CIBC OR Scotiabank e-transfer problem site:reddit.com",
        "e-transfer limit increase OR workaround site:reddit.com",
        "e-transfer scam send back site:reddit.com",
        "interac e-transfer delay hours site:reddit.com",
        "e-transfer not received site:reddit.com/r/personalfinancecanada",
        "e-transfer fraud unauthorized site:reddit.com",
    ]
    for q in reddit_ddg_fallback:
        rr = await web_search(q, "search", 8, tbs=ddg_tbs)
        for r in rr:
            link = r.get("link", "")
            if not link or link in seen_links or _is_blocked_domain(link):
                continue
            seen_links.add(link)
            r["channel"] = "people"
            r["source"] = "Reddit"
            r["_fetch_method"] = "ddg_reddit_fallback"
            etransfer_social.append(r)
    logger.info(f"[reddit-ddg-fallback] etransfer_social now {len(etransfer_social)} posts")

    # ── Platform helpers (defined here so tiered-fetch can use them) ──────────
    def _is_reddit(m: dict) -> bool:
        return "reddit.com" in (m.get("link") or "").lower() or m.get("source") == "Reddit"

    def _is_twitter(m: dict) -> bool:
        return m.get("source") == "X/Twitter"

    # ── Tiered fetch: decide whether to run DDG ──────────────────────────────
    # Require BOTH platforms to have content before skipping DDG.
    # If Reddit API is blocked on Railway, reddit_count=0 → DDG always runs,
    # restoring site:reddit.com fallback queries.
    reddit_count_pre  = sum(1 for m in etransfer_social if _is_reddit(m))
    twitter_count_pre = sum(1 for m in etransfer_social if _is_twitter(m))
    comp_threshold    = 5
    reddit_min        = 8
    twitter_min       = 8

    skip_ddg = (
        reddit_count_pre  >= reddit_min
        and twitter_count_pre >= twitter_min
        and len(competitor_mentions) >= comp_threshold
    )
    logger.info(
        f"[tiered-fetch] reddit={reddit_count_pre} twitter={twitter_count_pre} "
        f"competitor={len(competitor_mentions)} skip_ddg={skip_ddg}"
    )

    # ── 3b + 4. DDG + Google News — parallel fetch with semaphore ──────────────
    # Semaphore limits to 5 concurrent queries to avoid DDG rate-limiting.
    # Each query fetches sub-searches concurrently. Results are post-processed
    # in order so seen_links dedup is safe (no shared state during gather).
    if not skip_ddg:
        ddg_sem = asyncio.Semaphore(5)

        async def _fetch_et_query(query: str) -> dict:
            async with ddg_sem:
                tasks: list = [web_search(query, "search", 8, tbs=ddg_tbs)]
                if not _has_site_restriction(query):
                    tasks.append(search_google_news(query, max_results=8))
                    tasks.append(search_twitter(query, 4, tbs=ddg_tbs))
                res = await asyncio.gather(*tasks, return_exceptions=True)
            text = res[0] if not isinstance(res[0], Exception) else []
            news = (res[1] if not isinstance(res[1], Exception) else []) if len(res) > 1 else []
            twit = (res[2] if not isinstance(res[2], Exception) else []) if len(res) > 2 else []
            return {"text": text, "news": news, "twitter": twit}

        async def _fetch_comp_query(query: str) -> dict:
            async with ddg_sem:
                tasks: list = [web_search(query, "search", 12, tbs=ddg_tbs)]
                if not _has_site_restriction(query):
                    tasks.append(search_google_news(query, max_results=8))
                res = await asyncio.gather(*tasks, return_exceptions=True)
            text = res[0] if not isinstance(res[0], Exception) else []
            news = (res[1] if not isinstance(res[1], Exception) else []) if len(res) > 1 else []
            return {"text": text, "news": news}

        et_batches, comp_batches = await asyncio.gather(
            asyncio.gather(*[_fetch_et_query(q) for q in etransfer_ddg_queries], return_exceptions=True),
            asyncio.gather(*[_fetch_comp_query(q) for q in competitor_ddg_queries], return_exceptions=True),
        )

        # ── Post-process e-Transfer DDG results ──
        for item in et_batches:
            if isinstance(item, Exception):
                continue
            for r in item["text"]:
                link = r.get("link", "")
                if not link or link in seen_links or _is_blocked_domain(link):
                    continue
                seen_links.add(link)
                channel, source = _classify_channel_and_source(link)
                r["channel"] = channel
                r["source"] = source
                r["_fetch_method"] = "ddg_text"
                if channel == "people":
                    etransfer_social.append(r)
                else:
                    etransfer_press.append(r)
            for r in item["news"]:
                link = r.get("link", "")
                if not link or link in seen_links or _is_blocked_domain(link):
                    continue
                seen_links.add(link)
                channel, source = _classify_channel_and_source(link)
                r["channel"] = channel
                r["source"] = source
                if channel == "people":
                    etransfer_social.append(r)
                elif not _is_low_quality_market_content(r):
                    etransfer_press.append(r)
            for r in item["twitter"]:
                link = r.get("link", "")
                if not link or link in seen_links or _is_blocked_domain(link):
                    continue
                seen_links.add(link)
                r["channel"] = "people"
                r["source"] = "X/Twitter"
                r["_fetch_method"] = "ddg_text"
                etransfer_social.append(r)

        # ── Post-process competitor DDG results ──
        for item in comp_batches:
            if isinstance(item, Exception):
                continue
            for r in item["text"]:
                link = r.get("link", "")
                if not link or link in seen_links or _is_blocked_domain(link):
                    continue
                if _is_low_quality_market_content(r):
                    continue
                seen_links.add(link)
                channel, source = _classify_channel_and_source(link)
                r["channel"] = channel
                r["source"] = source
                r["_fetch_method"] = "ddg_search"
                competitor_mentions.append(r)
            for r in item["news"]:
                link = r.get("link", "")
                if not link or link in seen_links or _is_blocked_domain(link):
                    continue
                if _is_low_quality_market_content(r):
                    continue
                seen_links.add(link)
                r["channel"] = "press"
                r["_fetch_method"] = "google_news_rss"
                competitor_mentions.append(r)

    # Date backfill: Reddit .json for undated thread URLs (e.g. DDG-only Reddit hits),
    # then HTML <head> meta for other undated URLs (cap excludes Reddit threads filled above).
    all_mentions = etransfer_social + etransfer_press + competitor_mentions
    await _enrich_dates_from_reddit_json(all_mentions, max_fetches=40)
    await _enrich_dates_from_meta(all_mentions, max_fetches=80)

    # Score by quality, then stratify by platform to ensure both Reddit and Twitter
    # get representational slots before Kimi filter. Prevents Reddit volume from crowding
    # out Twitter in the top-50 cut. Quality ordering preserved within each platform.
    all_scored = sorted(
        [(m, _mention_quality_score(m, "etransfer")) for m in etransfer_social],
        key=lambda x: x[1], reverse=True,
    )

    # Stratify: take top N from each platform, then merge
    reddit_scored = [(m, s) for m, s in all_scored if _is_reddit(m)]
    twitter_scored = [(m, s) for m, s in all_scored if _is_twitter(m)]
    other_scored = [(m, s) for m, s in all_scored if not _is_reddit(m) and not _is_twitter(m)]

    if quarterly:
        reddit_pool = [m for m, _ in reddit_scored[:28]]
        twitter_pool = [m for m, _ in twitter_scored[:18]]
        other_pool = [m for m, _ in other_scored[:14]]
    else:
        reddit_pool = [m for m, _ in reddit_scored[:17]]
        twitter_pool = [m for m, _ in twitter_scored[:11]]
        other_pool = [m for m, _ in other_scored[:7]]

    etransfer_social = reddit_pool + twitter_pool + other_pool

    reddit_count  = sum(1 for m in etransfer_social if _is_reddit(m))
    twitter_count = sum(1 for m in etransfer_social if _is_twitter(m))
    other_count   = len(etransfer_social) - reddit_count - twitter_count
    logger.info(
        f"[quality-sort] pool={len(etransfer_social)} "
        f"(reddit={reddit_count}, twitter={twitter_count}, other={other_count})"
    )

    # Make recency decisions only on dated content to avoid stale/undated drift.
    etransfer_social = _filter_recent_dated_mentions(
        etransfer_social, max_age_days=max_age_days
    )
    etransfer_press = _filter_recent_dated_mentions(
        etransfer_press, max_age_days=max_age_days
    )
    competitor_mentions = _filter_recent_dated_mentions(
        competitor_mentions, max_age_days=max_age_days
    )

    # ── Source breakdown log ──────────────────────────────────────────────────
    def _src_counts(items: list[dict]) -> str:
        from collections import Counter
        c = Counter(m.get("_fetch_method", "unknown") for m in items)
        return " | ".join(f"{k}={v}" for k, v in sorted(c.items()))

    logger.info(
        f"[fetch-sources] BEFORE quality gate — "
        f"etransfer_social={len(etransfer_social)} ({_src_counts(etransfer_social)}) | "
        f"etransfer_press={len(etransfer_press)} ({_src_counts(etransfer_press)}) | "
        f"competitor={len(competitor_mentions)} ({_src_counts(competitor_mentions)})"
    )

    logger.info(
        f"[fetch-sources] after recency filter — "
        f"etransfer_social={len(etransfer_social)} ({_src_counts(etransfer_social)}) | "
        f"etransfer_press={len(etransfer_press)} ({_src_counts(etransfer_press)}) | "
        f"competitor={len(competitor_mentions)} ({_src_counts(competitor_mentions)})"
    )

    # ── Kimi value filter: the real quality gate ─────────────────────────────
    # Runs Kimi over every mention and scores it 1–5 on insight value.
    # Anything scoring <3 is dropped before the final generation prompt sees it.
    # This is what catches "Yes Interac e-Transfer. No fees 🇨🇦" regex can't.
    # Save pre-filter pool to allow diversity floor afterwards.
    etransfer_social_prefilter = list(etransfer_social)
    competitor_mentions_prefilter = list(competitor_mentions)
    # DDG-sourced Reddit posts are shorter/less specific — lower threshold to 2
    # so the diversity floor has real candidates to pull from.
    _ddg_reddit = [m for m in etransfer_social if m.get("_fetch_method") == "ddg_reddit_fallback"]
    _regular_et = [m for m in etransfer_social if m.get("_fetch_method") != "ddg_reddit_fallback"]
    logger.info(
        f"[value-filter] etransfer_social split — ddg_reddit={len(_ddg_reddit)} (min_score=2) | regular={len(_regular_et)} (min_score=3)"
    )
    logger.info("[value-filter] running Kimi value filter on 4 buckets in parallel...")
    _ddg_reddit_filtered, _regular_et_filtered, etransfer_press, competitor_mentions = await asyncio.gather(
        kimi_filter_by_value(_ddg_reddit, min_score=2),
        kimi_filter_by_value(_regular_et, min_score=3),
        kimi_filter_by_value(etransfer_press, min_score=3),
        kimi_filter_by_value(competitor_mentions, min_score=3),
    )
    etransfer_social = _regular_et_filtered + _ddg_reddit_filtered
    logger.info(
        f"[fetch-sources] after Kimi value filter — "
        f"etransfer_social={len(etransfer_social)} | "
        f"etransfer_press={len(etransfer_press)} | "
        f"competitor={len(competitor_mentions)}"
    )

    # ── Diversity floor (post-filter) ────────────────────────────────────────
    # If after all filtering one platform is completely absent but the other
    # has plenty, pull in the top 1–2 pre-filter items from the empty platform.
    # Quality stays primary — we only restore diversity when the gap is stark.
    def _platform_count(items: list[dict], check: callable) -> int:
        return sum(1 for m in items if check(m))

    reddit_after = _platform_count(etransfer_social, _is_reddit)
    twitter_after = _platform_count(etransfer_social, _is_twitter)

    if reddit_after < 2 and twitter_after >= 3:
        # Twitter dominates; ensure minimum Reddit representation (floor of 3)
        needed = 3 - reddit_after
        reddit_rescue = [m for m in etransfer_social_prefilter if _is_reddit(m)][:needed]
        if reddit_rescue:
            etransfer_social = etransfer_social + reddit_rescue
            logger.info(f"[diversity-floor] rescued {len(reddit_rescue)} Reddit items to floor of 3")

    if twitter_after < 2 and reddit_after >= 3:
        # Reddit dominates; ensure minimum Twitter representation (floor of 3)
        needed = 3 - twitter_after
        twitter_rescue = [m for m in etransfer_social_prefilter if _is_twitter(m)][:needed]
        if twitter_rescue:
            etransfer_social = etransfer_social + twitter_rescue
            logger.info(f"[diversity-floor] rescued {len(twitter_rescue)} Twitter items to floor of 3")

    # ── Market Pulse floor ──────────────────────────────────────────────────────
    # If competitor mentions drop below 3 after Kimi filter, pull highest-scoring
    # pre-filter items (which may have score-2) up to a floor of 4.
    if len(competitor_mentions) < 3:
        # Get pre-filter items, sorted by quality score, take top N to reach floor
        competitor_rescue = sorted(
            [(m, _mention_quality_score(m, "competitor")) for m in competitor_mentions_prefilter],
            key=lambda x: x[1], reverse=True
        )
        needed = 4 - len(competitor_mentions)
        rescue_items = [m for m, _ in competitor_rescue[:needed]]
        if rescue_items:
            competitor_mentions = competitor_mentions + rescue_items
            logger.info(f"[market-floor] rescued {len(rescue_items)} competitor items to reach floor of 4")

    total = len(etransfer_social) + len(etransfer_press) + len(competitor_mentions)
    if total == 0:
        if quarterly:
            last_quarterly_url_dates = {}
        else:
            last_biweekly_url_dates = {}
        return ("No mentions found in the configured lookback window.", [])

    scan_banner = "QUARTERLY" if quarterly else "BIWEEKLY"
    lines = [f"=== INTERAC {scan_banner} SCAN — {now_est()} ==="]
    lines.append(
        f"Total: {total} mentions "
        f"({len(etransfer_social)} e-Transfer community, {len(etransfer_press)} e-Transfer news, "
        f"{len(competitor_mentions)} competitor)"
    )
    lines.append("")

    def _fmt(prefix: str, items: list[dict], cap: int, snippet_cap: int) -> list[str]:
        out = []
        for i, m in enumerate(items[:cap], 1):
            snippet = " ".join((m.get("snippet", "") or "").split())[:snippet_cap]
            out.append(f"[{prefix}{i}] {m.get('source', 'unknown')}")
            # Explicit Date: field on its own line so the LLM reliably picks it up
            out.append(f"  Date: {m['date']}" if m.get("date") else "  Date: unknown")
            out.append(f"  Title: {m.get('title', '')[:120]}")
            out.append(f"  Snippet: {snippet}")
            out.append(f"  URL: {m.get('link', '')}")
            out.append("")
        return out

    s_cap, en_cap, c_cap = (100, 28, 100) if quarterly else (60, 15, 60)
    s_snip, en_snip, c_snip = (900, 400, 550) if quarterly else (750, 300, 450)

    if etransfer_social:
        lines.append("=== e-TRANSFER COMMUNITY (REDDIT, RFD, X) ===")
        lines += _fmt("S", etransfer_social, s_cap, s_snip)

    if etransfer_press:
        lines.append("=== e-TRANSFER NEWS ===")
        lines += _fmt("EN", etransfer_press, en_cap, en_snip)

    if competitor_mentions:
        lines.append("=== COMPETITOR INTELLIGENCE (Wise, PayPal, Apple Pay, Wealthsimple, KOHO, Venmo, Zelle, Revolut, Neo, ACH, others) ===")
        lines += _fmt("C", competitor_mentions, c_cap, c_snip)

    url_map = _build_url_date_map_from_mentions(
        etransfer_social, etransfer_press, competitor_mentions
    )
    if quarterly:
        last_quarterly_url_dates = url_map
    else:
        last_biweekly_url_dates = url_map

    def _mention_source_row(m: dict, *, bucket: str, snippet_cap: int) -> dict:
        link = (m.get("link") or "").strip()
        ch, _src = _classify_channel_and_source(link)
        sec = "competitor" if bucket == "competitor_intelligence" else "etransfer"
        snip = " ".join((m.get("snippet", "") or "").split())[:snippet_cap]
        return {
            "source_bucket": bucket,
            "url_original": link,
            "source_label": m.get("source", "") or "",
            "channel": (m.get("channel") or ch or ""),
            "published_date": m.get("date") or "",
            "title": (m.get("title", "") or "")[:120],
            "snippet_included_in_prompt": snip,
            "quality_score_heuristic": round(float(_mention_quality_score(m, sec)), 3),
        }

    sources: list[dict] = []
    seen_pair: set[tuple[str, str]] = set()

    def _add_source_row(m: dict, *, bucket: str, snippet_cap: int) -> None:
        row = _mention_source_row(m, bucket=bucket, snippet_cap=snippet_cap)
        key = (row["url_original"], row["source_bucket"])
        if not row["url_original"] or key in seen_pair:
            return
        seen_pair.add(key)
        sources.append(row)

    for m in etransfer_social[:s_cap]:
        _add_source_row(m, bucket="e_transfer_community", snippet_cap=s_snip)
    for m in etransfer_press[:en_cap]:
        _add_source_row(m, bucket="e_transfer_news", snippet_cap=en_snip)
    for m in competitor_mentions[:c_cap]:
        _add_source_row(m, bucket="competitor_intelligence", snippet_cap=c_snip)

    return ("\n".join(lines), sources)


# ─── Kimi K2.5 Analysis ──────────────────────────────────────────────────────
def _kimi_max_user_chars_for_context(
    system_prompt: str, *, max_tokens: int, max_user_chars: int, margin: int = 220
) -> int:
    """Moonshot chat completions enforce a total context budget (often 8192 on kimi-k2.5-preview).

    Reserve space for system prompt, requested completion, and framing. Returns a safe
    user-message character cap (conservative ~3 chars per token for mixed text).
    """
    try:
        max_ctx = int(os.environ.get("KIMI_MAX_CONTEXT_TOKENS", "8192"))
    except ValueError:
        max_ctx = 8192
    # Rough token estimate from UTF-8 text (good enough for budgeting).
    sys_tok = max(1, (len(system_prompt) + 3) // 4)
    reserved = sys_tok + max_tokens + margin
    room = max_ctx - reserved
    if room < 256:
        room = 256
    by_ctx = room * 3
    return min(max_user_chars, by_ctx)


async def call_kimi(
    system_prompt: str,
    user_content: str,
    *,
    max_user_chars: int = 10000,
    max_tokens: int = 2000,
    temperature: float = 0.3,
    timeout_sec: float = 120.0,
) -> str:
    # Stay under provider total-token limit (system + user + max_tokens).
    cap = _kimi_max_user_chars_for_context(
        system_prompt, max_tokens=max_tokens, max_user_chars=max_user_chars
    )
    if len(user_content) > cap:
        head = (cap * 2) // 3
        tail = cap - head - 90
        if tail < 256:
            user_content = user_content[:cap] + "\n\n[... truncated for API context limit ...]"
        else:
            user_content = (
                user_content[:head]
                + "\n\n[... truncated for API context limit ...]\n\n"
                + user_content[-tail:]
            )

    async with httpx.AsyncClient(timeout=timeout_sec) as client:
        response = await client.post(
            KIMI_API_URL,
            headers={
                "Authorization": f"Bearer {KIMI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": KIMI_MODEL,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],
                "temperature": temperature,
                "max_tokens": max_tokens,
            },
        )
        if response.status_code != 200:
            body = response.text
            logger.error(f"Kimi API {response.status_code}: {body}")
            raise Exception(f"Kimi API {response.status_code}: {body[:300]}")
        data = response.json()
        return data["choices"][0]["message"]["content"]


_VALUE_FILTER_SYSTEM = """You are a strict quality gatekeeper for an Interac e-Transfer product intelligence report. Your job is to score each numbered item 1–5 on insight value for a payments product manager.

# SCORING RUBRIC

**5 — Excellent (include):** Specific personal experience with concrete details (dollar amount, bank name, timeframe) AND a clear problem/frustration/comparison/workaround.
Example: "TD held my $2,400 e-Transfer for 5 days, their agent blamed Interac."

**4 — Strong (include):** Personal experience OR a concrete market development with specific detail. Clear insight.
Example: "Switched from e-Transfer to Wise for anything over $1k — saves $40 per transfer."

**3 — Borderline (include):** Has some specific content — a named product, a specific comparison, or a real user question with context. MUST have either personal experience ("I", "my", "we") OR a concrete event/development. Pure editorial opinion or analysis — even if it contains industry statistics — is NOT score 3.
Example: "Does anyone else's bank charge for Interac e-Transfer now?"
NOT score 3: "Interac is a monopoly, they charge $0.20/tx" — this is opinion/analysis, score 2.

**2 — Weak (EXCLUDE):** Generic mention, no personal experience, no concrete detail. IMPORTANT: Opinion or editorial content that contains statistics or industry data (e.g. "Interac is a monopoly, they charge $0.20/tx and processed 2 billion transactions") scores 2, NOT 3 — it is analysis/opinion, not a personal user experience. Planning or future-tense posts ("I plan to make a transfer today") score 2.
Example: "e-Transfer is convenient for sending money."
Example: "Interac is a monopoly because they control all P2P payments in Canada."
Example: "I'm going to test if I get prompted for 2FA next time."

**1 — Junk (EXCLUDE):** Promotional, explainer content, crypto/web3 tangents, affirmations with no insight, prize/contest mentions, off-topic.
Examples:
- "Yes Interac e-Transfer from your Canadian bank. No fees 🇨🇦"
- "e-transfer is a product from Interac, which is an interbank network..."
- "This @daimo integration delivers a clean off-ramp from @CADCstablecoin via Interac e-Transfer"
- "Win $500 via e-Transfer!"

# OUTPUT FORMAT

For EVERY numbered item you receive, output exactly one line:

`<number>|<score>|<one-line reason>`

Example output:
```
1|5|specific TD hold with dollar amount
2|1|crypto stablecoin tangent, no user experience
3|2|generic affirmation, no detail
4|4|clear Wise vs e-Transfer comparison
```

Output scores for ALL items. Nothing else. No preamble, no summary."""


def _build_value_filter_input(mentions: list[dict]) -> str:
    """Format mentions as numbered blocks for Kimi scoring."""
    lines = []
    for i, m in enumerate(mentions, 1):
        snippet = " ".join((m.get("snippet", "") or "").split())[:400]
        title = (m.get("title", "") or "")[:120]
        source = m.get("source", "unknown")
        lines.append(f"[{i}] {source}")
        lines.append(f"  Title: {title}")
        lines.append(f"  Snippet: {snippet}")
        lines.append("")
    return "\n".join(lines)


def _parse_value_scores(response: str, count: int) -> dict[int, int]:
    """Parse '<n>|<score>|<reason>' lines from Kimi into {index: score}."""
    scores = {}
    for line in response.splitlines():
        line = line.strip()
        if not line or "|" not in line:
            continue
        parts = line.split("|", 2)
        if len(parts) < 2:
            continue
        try:
            idx = int(parts[0].strip().lstrip("[").rstrip("]"))
            score = int(parts[1].strip())
            if 1 <= idx <= count and 1 <= score <= 5:
                scores[idx] = score
        except ValueError:
            continue
    return scores


async def kimi_filter_by_value(mentions: list[dict], min_score: int = 3) -> list[dict]:
    """Run Kimi value filter. Returns only mentions scoring >= min_score.

    Items not scored by Kimi (parse failures) are kept as a safety fallback —
    better to let the final prompt see them than silently drop them.
    """
    if not mentions or not KIMI_API_KEY:
        return mentions

    user_content = _build_value_filter_input(mentions)
    if len(user_content) > 18000:
        user_content = user_content[:18000]

    # Output is "N|score|reason" per item — ~40 chars × 50 items ≈ 600 tokens.
    # Cap at 1200 max_tokens: enough headroom, much faster than 3000.
    try:
        async with httpx.AsyncClient(timeout=45) as client:
            r = await client.post(
                KIMI_API_URL,
                headers={"Authorization": f"Bearer {KIMI_API_KEY}", "Content-Type": "application/json"},
                json={
                    "model": KIMI_MODEL,
                    "messages": [
                        {"role": "system", "content": _VALUE_FILTER_SYSTEM},
                        {"role": "user", "content": user_content},
                    ],
                    "temperature": 0.1,
                    "max_tokens": 1200,
                },
            )
            if r.status_code != 200:
                logger.warning(f"[value-filter] Kimi {r.status_code}: {r.text[:200]}")
                return mentions  # fallback: keep everything
            raw = r.json()["choices"][0]["message"]["content"]
    except Exception as e:
        logger.warning(f"[value-filter] error: {e} — skipping filter")
        return mentions

    scores = _parse_value_scores(raw, len(mentions))
    kept = []
    drop_scores = []
    for i, m in enumerate(mentions, 1):
        score = scores.get(i)
        if score is None:
            kept.append(m)  # unscored → keep (safe default)
        elif score >= min_score:
            m["_kimi_value"] = score
            kept.append(m)
        else:
            drop_scores.append(score)

    logger.info(
        f"[value-filter] {len(mentions)} in → {len(kept)} kept "
        f"(dropped {len(drop_scores)}, scored={len(scores)}, unscored={len(mentions) - len(scores)})"
    )
    return kept


async def curate_with_kimi(raw_mentions: str) -> str:
    """Kimi curation pass: filter raw mentions down to real-people social signal before analysis."""
    config = load_prompts()
    prompt = config.get("curation_prompt", "")
    if not prompt:
        logger.warning("[curation] curation_prompt not found — skipping curation step")
        return raw_mentions

    user_content = raw_mentions
    if len(user_content) > 10000:
        user_content = user_content[:10000] + "\n\n[... truncated]"

    async with httpx.AsyncClient(timeout=120) as client:
        response = await client.post(
            KIMI_API_URL,
            headers={
                "Authorization": f"Bearer {KIMI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": KIMI_MODEL,
                "messages": [
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": user_content},
                ],
                "temperature": 0.3,
                "max_tokens": 1500,
            },
        )
        if response.status_code != 200:
            body = response.text
            logger.error(f"[curation] Kimi API {response.status_code}: {body[:300]}")
            return raw_mentions  # graceful fallback: use unfiltered data
        data = response.json()
        curated = data["choices"][0]["message"]["content"]
        logger.info(
            f"[curation] done — input {len(raw_mentions)} chars → output {len(curated)} chars"
        )
        return curated


def _split_mentions_sections(mentions_text: str) -> tuple[str, str]:
    """Split fetch output into (community_text, market_text).
    community_text = e-TRANSFER COMMUNITY only (left side).
    market_text = e-TRANSFER NEWS + COMPETITOR INTELLIGENCE (right side).
    """
    news_marker = "=== e-TRANSFER NEWS"
    comp_marker = "=== COMPETITOR INTELLIGENCE"
    community_text = mentions_text
    market_text = ""
    for marker in [news_marker, comp_marker]:
        if marker in community_text:
            idx = community_text.index(marker)
            market_text = community_text[idx:] + ("\n\n" + market_text if market_text else "")
            community_text = community_text[:idx]
    return community_text.strip(), market_text.strip()


def _normalize_url_for_match(url: str) -> str:
    """Normalize URL for joining report bullets to ledger rows (not persisted as its own column)."""
    u = (url or "").strip().rstrip(").,;]")
    if not u:
        return ""
    if not u.startswith("http"):
        u = "https://" + u
    try:
        p = urlparse(u)
        netloc = (p.netloc or "").lower()
        if netloc.startswith("www."):
            netloc = netloc[4:]
        path = (p.path or "").rstrip("/").lower()
        return f"{netloc}{path}"
    except Exception:
        return re.sub(r"\?.*$", "", u.lower())


def _norm_url_to_bullets(section_text: str) -> dict[str, list[str]]:
    """Map normalized URL -> list of bullet lines in that section containing the URL."""
    out: dict[str, list[str]] = {}
    for line in (section_text or "").splitlines():
        s = line.strip()
        if "http" not in s:
            continue
        if not (s.startswith("-") or s.startswith("•") or re.match(r"^\d+\.\s", s)):
            continue
        for raw_u in re.findall(r"https?://[^\s\)\]<>\"']+", s):
            key = _normalize_url_for_match(raw_u)
            if key:
                out.setdefault(key, []).append(s)
    return out


_CHATTER_KIMI_TAG_RE = re.compile(
    r"^\[\s*(Praise|Comparison|Education|Blame|Thin\s+mention)\s*\]\s+",
    re.IGNORECASE,
)


def _split_chatter_kimi_tag(text: str) -> tuple[str, str]:
    """Parse leading ``[Praise]``-style tag. Returns ``(category_key, rest)``; key empty if missing."""
    s = (text or "").strip()
    m = _CHATTER_KIMI_TAG_RE.match(s)
    if not m:
        return ("", s)
    raw = re.sub(r"\s+", " ", m.group(1).strip().lower())
    key_map = {
        "praise": "praise",
        "comparison": "comparison",
        "education": "education",
        "blame": "blame",
        "thin mention": "thin",
    }
    key = key_map.get(raw, "")
    return (key, s[m.end() :])


def _chatter_category_from_line(stripped: str) -> str:
    """Kimi ``[Tag]`` on the bullet if present, else keyword fallback (same keys as mix bars)."""
    st = stripped.strip()
    if not (st.startswith("- ") or st.startswith("• ")):
        return "thin"
    body = st[2:].strip()
    tag_key, _ = _split_chatter_kimi_tag(body)
    if tag_key:
        return tag_key
    return _classify_chatter_bullet_line(stripped)


def _chatter_body_after_tag(stripped: str) -> str:
    """Content after ``- `` / ``• `` with optional ``[Tag] `` removed for quote vs attribution split."""
    st = stripped.strip()
    if not (st.startswith("- ") or st.startswith("• ")):
        return st.strip()
    body = st[2:].strip()
    tag_key, rest = _split_chatter_kimi_tag(body)
    return rest if tag_key else body


def _norm_url_to_chatter_lines(
    section_text: str,
) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    """Like ``_norm_url_to_bullets`` for e-Transfer Chatter, plus parallel Kimi category keys per URL."""
    lines_out: dict[str, list[str]] = {}
    cats_out: dict[str, list[str]] = {}
    for line in (section_text or "").splitlines():
        s = line.strip()
        if "http" not in s:
            continue
        if not (s.startswith("-") or s.startswith("•") or re.match(r"^\d+\.\s", s)):
            continue
        cat = _chatter_category_from_line(s)
        for raw_u in re.findall(r"https?://[^\s\)\]<>\"']+", s):
            key = _normalize_url_for_match(raw_u)
            if key:
                lines_out.setdefault(key, []).append(s)
                cats_out.setdefault(key, []).append(cat)
    return lines_out, cats_out


def _norm_url_to_lines_any(text: str) -> dict[str, list[str]]:
    """Like _norm_url_to_bullets but matches any line containing a URL (for quarterly prose)."""
    out: dict[str, list[str]] = {}
    for line in (text or "").splitlines():
        if "http" not in line:
            continue
        for raw_u in re.findall(r"https?://[^\s\)\]<>\"']+", line):
            key = _normalize_url_for_match(raw_u)
            if key:
                out.setdefault(key, []).append(line.strip())
    return out


def _excerpt_around_url(text: str, url_norm: str, *, radius: int = 450) -> str:
    if not text or not url_norm:
        return ""
    lower = text.lower()
    idx = -1
    for m in re.finditer(r"https?://[^\s\)\]<>\"']+", text):
        if _normalize_url_for_match(m.group(0)) == url_norm:
            idx = m.start()
            break
    if idx < 0:
        return ""
    start = max(0, idx - radius)
    end = min(len(text), idx + radius)
    return " ".join(text[start:end].split())


def _append_source_ledger(
    *,
    run_type: str,
    report_scan_datetime: str,
    sources: list[dict],
    biweekly_report_for_match: str | None = None,
    quarterly_report_for_match: str | None = None,
    quarterly_digest: str | None = None,
    quarterly_digest_used: bool | None = None,
) -> None:
    """Append one row per source to state/source_ledger.xlsx (see plan for columns)."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        logger.warning("openpyxl not available — skipping source ledger")
        return

    run_cal = datetime.now(EST).date().isoformat()
    chatter_map: dict[str, list[str]] = {}
    chatter_cat_map: dict[str, list[str]] = {}
    market_map: dict[str, list[str]] = {}
    if biweekly_report_for_match:
        chatter_raw = _extract_section(
            biweekly_report_for_match,
            "e-Transfer Chatter:",
            ["Market Pulse:", "Trend vs Last Scan:"],
        )
        market_raw = _extract_section(
            biweekly_report_for_match,
            "Market Pulse:",
            ["Trend vs Last Scan:"],
        )
        chatter_map, chatter_cat_map = _norm_url_to_chatter_lines(chatter_raw)
        market_map = _norm_url_to_bullets(market_raw)

    q_report_map: dict[str, list[str]] = {}
    if quarterly_report_for_match:
        q_report_map = _norm_url_to_lines_any(quarterly_report_for_match)

    q_digest_map: dict[str, list[str]] = {}
    if quarterly_digest:
        q_digest_map = _norm_url_to_lines_any(quarterly_digest)

    written_at = datetime.now(timezone.utc).isoformat()

    header = [
        "run_type",
        "report_scan_datetime",
        "run_calendar_date",
        "source_bucket",
        "url_original",
        "source_label",
        "channel",
        "published_date",
        "title",
        "snippet_included_in_prompt",
        "quality_score_heuristic",
        "in_biweekly_chatter",
        "in_biweekly_market_pulse",
        "biweekly_chatter_bullet",
        "biweekly_chatter_category",
        "biweekly_market_bullet",
        "in_quarterly_final_report",
        "quarterly_report_excerpt",
        "in_quarterly_compress_digest",
        "quarterly_digest_excerpt",
        "ledger_written_at",
        "notes",
    ]

    SOURCE_LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
    if SOURCE_LEDGER_PATH.exists():
        wb = load_workbook(SOURCE_LEDGER_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sources"
        ws.append(header)

    def _join_bullets(xs: list[str] | None) -> str:
        if not xs:
            return ""
        return " | ".join(xs)

    def _join_chatter_categories(keys: list[str] | None) -> str:
        if not keys:
            return ""
        return " | ".join(CHATTER_CAT_LABELS.get(k, k) for k in keys)

    hdr_row = [c.value for c in ws[1]] if ws.max_row else []
    if hdr_row and "biweekly_chatter_category" not in hdr_row and "biweekly_chatter_bullet" in hdr_row:
        j = hdr_row.index("biweekly_chatter_bullet")
        insert_at = j + 2
        ws.insert_cols(insert_at, amount=1)
        ws.cell(row=1, column=insert_at, value="biweekly_chatter_category")
        hdr_row = [c.value for c in ws[1]]

    for src in sources:
        url_o = src.get("url_original") or ""
        nk = _normalize_url_for_match(url_o)
        notes = ""

        if run_type == "biweekly":
            in_ch = "yes" if nk and nk in chatter_map else "no"
            in_mp = "yes" if nk and nk in market_map else "no"
            b_ch = _join_bullets(chatter_map.get(nk))
            b_ch_cat = _join_chatter_categories(chatter_cat_map.get(nk))
            b_mp = _join_bullets(market_map.get(nk))
            in_q = "n_a"
            q_ex = ""
            in_qd = "n_a"
            qd_ex = ""
        else:
            in_ch = "n_a"
            in_mp = "n_a"
            b_ch = ""
            b_ch_cat = ""
            b_mp = ""
            in_q = "yes" if nk and nk in q_report_map else "no"
            q_ex = _excerpt_around_url(quarterly_report_for_match or "", nk) if quarterly_report_for_match else ""
            if quarterly_digest_used is True:
                in_qd = "yes" if nk and nk in q_digest_map else "no"
                qd_ex = _excerpt_around_url(quarterly_digest or "", nk) if quarterly_digest else ""
            elif quarterly_digest_used is False:
                in_qd = "n_a"
                qd_ex = ""
            else:
                in_qd = "n_a"
                qd_ex = ""

        if nk and in_ch == "yes" and in_mp == "yes":
            notes = "matched_both_biweekly_columns"

        row_vals = {
            "run_type": run_type,
            "report_scan_datetime": report_scan_datetime,
            "run_calendar_date": run_cal,
            "source_bucket": src.get("source_bucket", ""),
            "url_original": url_o,
            "source_label": src.get("source_label", ""),
            "channel": src.get("channel", ""),
            "published_date": src.get("published_date", ""),
            "title": src.get("title", ""),
            "snippet_included_in_prompt": src.get("snippet_included_in_prompt", ""),
            "quality_score_heuristic": src.get("quality_score_heuristic", ""),
            "in_biweekly_chatter": in_ch,
            "in_biweekly_market_pulse": in_mp,
            "biweekly_chatter_bullet": b_ch,
            "biweekly_chatter_category": b_ch_cat,
            "biweekly_market_bullet": b_mp,
            "in_quarterly_final_report": in_q,
            "quarterly_report_excerpt": q_ex,
            "in_quarterly_compress_digest": in_qd,
            "quarterly_digest_excerpt": qd_ex,
            "ledger_written_at": written_at,
            "notes": notes,
        }
        hdr = [c.value for c in ws[1]]
        ws.append([row_vals.get(h, "") for h in hdr])

    wb.save(SOURCE_LEDGER_PATH)
    logger.info(f"Appended {len(sources)} source ledger row(s) to {SOURCE_LEDGER_PATH}")


async def analyze_biweekly(mentions_text: str, sources: list[dict]) -> str:
    """Two independent Kimi calls — one per column — then combine into one report."""
    config = load_prompts()
    community_text, market_text = _split_mentions_sections(mentions_text)

    chatter_prompt = config.get("etransfer_chatter_prompt", "")
    market_prompt = config.get("market_pulse_prompt", "")

    # Run both Kimi calls in parallel
    chatter_task = asyncio.create_task(call_kimi(chatter_prompt, community_text)) if chatter_prompt else None
    market_task = asyncio.create_task(call_kimi(market_prompt, market_text)) if market_prompt and market_text else None

    chatter_bullets = await chatter_task if chatter_task else "Nothing notable this scan."
    market_bullets = await market_task if market_task else "Nothing notable this scan."
    chatter_bullets = _strip_model_ledger_lines(chatter_bullets)
    market_bullets = _strip_model_ledger_lines(market_bullets)

    scan_date = now_est()
    report_core = (
        f"SCAN DATE: {scan_date}\n\n"
        f"e-Transfer Chatter:\n{chatter_bullets.strip()}\n\n"
        f"Market Pulse:\n{market_bullets.strip()}\n\n"
        f"Trend vs Last Scan:\n"
        f"- Still active: none identified\n"
        f"- Went quiet: none identified\n"
        f"- New this scan: none identified"
    )

    themes = _extract_biweekly_themes(report_core)
    # Extract URLs from the full mentions pool so they're deduped next run
    sent_urls = re.findall(r"URL:\s*(https?://\S+)", mentions_text)
    _save_biweekly_memory(themes, scan_date, new_urls=sent_urls)
    footer = _source_ledger_footer()
    delivered = report_core + footer
    _append_biweekly_excel(scan_date, delivered)
    _append_source_ledger(
        run_type="biweekly",
        report_scan_datetime=scan_date,
        sources=sources,
        biweekly_report_for_match=report_core,
    )

    return delivered


# Quarterly analysis uses a map-reduce pattern so each Kimi request stays under typical
# 8192-token chat completion limits (system + user + max_tokens).
_QUARTERLY_CHUNK_CHARS = 3000
_QUARTERLY_DIGEST_JOIN_MAX = 12000
_QUARTERLY_COMPRESS_SYSTEM = """You compress raw Interac / competitor intelligence scan text into evidence lines.

Rules:
- Use ONLY the chunk below. Do not invent URLs, dates, or quotes.
- Output markdown bullets only. Each bullet: a short verbatim quote or tight paraphrase, then " — " + platform/source if known from the chunk, then " Source: URL" using a URL that appears in the chunk for that item.
- Prefer Interac e-Transfer, Canadian banks, and named payment competitors (Wise, PayPal, Wealthsimple, KOHO, Revolut, Apple Pay, Google Pay, Zelle, Venmo, Stripe, Square, Neo, etc.).
- Tag [Retail] or [Commercial] at the start of a bullet only when the speaker context is obvious (personal P2P vs business/payroll/vendor).
- If nothing here is usable, output exactly: (no signal in this chunk)
- Keep the entire response under 3200 characters."""


def _split_mentions_for_quarterly_compress(text: str, max_chars: int) -> list[str]:
    """Split scan text into newline-friendly chunks each under ~max_chars."""
    text = text.strip()
    if not text:
        return []
    if len(text) <= max_chars:
        return [text]
    blocks = [b.strip() for b in text.split("\n\n") if b.strip()]
    chunks: list[str] = []
    cur: list[str] = []
    cur_len = 0
    for b in blocks:
        extra = len(b) + (2 if cur else 0)
        if cur_len + extra > max_chars and cur:
            chunks.append("\n\n".join(cur))
            cur = [b]
            cur_len = len(b)
        else:
            cur.append(b)
            cur_len += extra
    if cur:
        chunks.append("\n\n".join(cur))
    out: list[str] = []
    for ch in chunks:
        if len(ch) <= max_chars:
            out.append(ch)
        else:
            for i in range(0, len(ch), max_chars):
                out.append(ch[i : i + max_chars])
    return out


async def _build_quarterly_evidence_digest(mentions_text: str) -> tuple[str, str]:
    """Return (full_merged_digest, digest_for_final_kimi_input — may be truncated for context)."""
    chunks = _split_mentions_for_quarterly_compress(mentions_text, _QUARTERLY_CHUNK_CHARS)
    if not chunks:
        return "", ""
    sem = asyncio.Semaphore(3)

    async def _compress_one(idx: int, body: str) -> str:
        async with sem:
            header = f"CHUNK {idx + 1} / {len(chunks)}\n\n"
            payload = header + body
            return await call_kimi(
                _QUARTERLY_COMPRESS_SYSTEM,
                payload,
                max_user_chars=len(payload) + 20,
                max_tokens=1100,
                temperature=0.15,
                timeout_sec=90.0,
            )

    parts = await asyncio.gather(*[_compress_one(i, c) for i, c in enumerate(chunks)])
    merged = "\n\n=== NEXT CHUNK ===\n\n".join(p.strip() for p in parts if p.strip())
    if len(merged) <= _QUARTERLY_DIGEST_JOIN_MAX:
        return merged, merged
    half = _QUARTERLY_DIGEST_JOIN_MAX // 2
    trimmed = (
        merged[:half]
        + "\n\n[... digest truncated: middle omitted for API context limit ...]\n\n"
        + merged[-half:]
    )
    return merged, trimmed


async def analyze_quarterly(mentions_text: str, sources: list[dict]) -> tuple[str, str | None, bool]:
    """Quarterly report via compress (map) + full prompt (reduce). Returns (delivered_text, digest_full_or_none, digest_map_reduce_used)."""
    config = load_prompts()
    prompt = (config.get("quarterly_market_trends_prompt") or "").strip()
    if not prompt:
        raise ValueError("quarterly_market_trends_prompt is missing from prompts.json / prompts folder")

    scan_date = now_est()
    prompt = prompt.replace("{timestamp}", scan_date)

    footer = _source_ledger_footer()

    # Fast path when the raw scan already fits one completion (short tests / thin pools).
    single_cap = _kimi_max_user_chars_for_context(
        prompt, max_tokens=2000, max_user_chars=20000
    )
    if len(mentions_text) + 400 <= single_cap:
        logger.info("[quarterly] single-pass analysis (raw fits API context budget)")
        report_core = _strip_model_ledger_lines(
            (
                await call_kimi(
                    prompt,
                    mentions_text,
                    max_user_chars=20000,
                    max_tokens=2000,
                    temperature=0.35,
                    timeout_sec=180.0,
                )
            ).strip()
        )
        delivered = report_core + footer
        _append_source_ledger(
            run_type="quarterly",
            report_scan_datetime=scan_date,
            sources=sources,
            quarterly_report_for_match=report_core,
            quarterly_digest=None,
            quarterly_digest_used=False,
        )
        return delivered, None, False

    digest_full, digest_for_model = await _build_quarterly_evidence_digest(mentions_text)
    if not digest_for_model.strip():
        digest_for_model = mentions_text[: _QUARTERLY_CHUNK_CHARS] + "\n\n[... raw truncated; compress pass empty ...]"
        digest_full = digest_for_model

    intro = (
        "The text below is a **pre-compressed evidence digest** built from the full quarterly "
        "raw scrape (split into API-sized chunks). Treat it as your only source material; "
        "do not invent facts that are not supported by it.\n\n"
    )
    user_block = intro + digest_for_model

    logger.info(
        f"[quarterly] final Kimi call — digest {len(digest_for_model)} chars, system {len(prompt)} chars"
    )
    report_core = _strip_model_ledger_lines(
        (
            await call_kimi(
                prompt,
                user_block,
                max_user_chars=20000,
                max_tokens=2000,
                temperature=0.35,
                timeout_sec=180.0,
            )
        ).strip()
    )
    delivered = report_core + footer
    _append_source_ledger(
        run_type="quarterly",
        report_scan_datetime=scan_date,
        sources=sources,
        quarterly_report_for_match=report_core,
        quarterly_digest=digest_full,
        quarterly_digest_used=True,
    )
    return delivered, digest_full, True


async def run_biweekly_scan(update: Update) -> None:
    """Run the universal biweekly e-Transfer intelligence scan and deliver to Telegram."""
    global last_report, last_mentions_raw
    tracked = _track_current_task()
    try:
        await update.message.reply_text(
            "Running biweekly e-Transfer intelligence scan (Reddit, X, RedFlagDeals, news)..."
        )
        mentions, sources = await asyncio.wait_for(fetch_biweekly_mentions(), timeout=300)
        last_mentions_raw = mentions

        if mentions.startswith("No mentions"):
            await update.message.reply_text(f"No data found this scan.\n\n{mentions[:500]}")
            return

        await update.message.reply_text("Mentions collected. Running Kimi curation pass, then analysis...")
        report = await asyncio.wait_for(analyze_biweekly(mentions, sources), timeout=180)
        last_report = report

        await send_chunked_message(
            update,
            f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}",
        )
    finally:
        _untrack_task(tracked)


def parse_email_modes() -> set[str]:
    # Supports: alert, weekly, always, comma-separated combinations.
    modes = {m.strip().lower() for m in EMAIL_SEND_MODE.split(",") if m.strip()}
    if not modes:
        modes = {"alert"}
    if "always" in modes:
        modes.update({"alert", "weekly"})
    return modes


WEEKDAY_TO_INDEX = {
    "monday": 0,
    "tuesday": 1,
    "wednesday": 2,
    "thursday": 3,
    "friday": 4,
    "saturday": 5,
    "sunday": 6,
}


def weekly_key(now_local: datetime) -> str:
    year, week_num, _ = now_local.isocalendar()
    return f"{year}-W{week_num}-{EMAIL_WEEKLY_DAY}-{EMAIL_WEEKLY_HOUR}"


def _should_send_email(
    *,
    trigger: str,
    now_local: datetime | None = None,
) -> tuple[bool, str]:
    """
    Decide whether we should send an email for this scan.
    Returns (should_send, reason).
    """
    if not EMAIL_ENABLED:
        return False, "EMAIL_ENABLED=0"

    modes = parse_email_modes()

    if trigger == "weekly" and "weekly" not in modes:
        return False, f"mode excludes weekly ({EMAIL_SEND_MODE})"

    if trigger == "weekly" and EMAIL_ALERT_DEDUP and now_local is not None:
        current_weekly_key = weekly_key(now_local)
        if current_weekly_key == last_weekly_email_key:
            return False, "weekly dedup"

    if trigger == "quarterly" and EMAIL_ALERT_DEDUP and now_local is not None:
        q_key = f"q-{now_local.date().isoformat()}"
        if q_key == last_quarterly_email_key:
            return False, "quarterly email dedup"

    if EMAIL_COOLDOWN_MINUTES > 0 and last_email_sent_at is not None:
        minutes_since = (datetime.now(timezone.utc) - last_email_sent_at).total_seconds() / 60.0
        if minutes_since < EMAIL_COOLDOWN_MINUTES:
            return False, f"cooldown {minutes_since:.1f}m/{EMAIL_COOLDOWN_MINUTES}m"

    return True, "ok"


def _smtp_config_summary() -> str:
    recipient_count = len(EMAIL_TO)
    user_hint = SMTP_USERNAME if SMTP_USERNAME else "(empty)"
    return (
        f"host={SMTP_HOST or '(empty)'} port={SMTP_PORT} "
        f"user={user_hint} from={EMAIL_FROM or '(empty)'} recipients={recipient_count}"
    )


def _resend_config_summary() -> str:
    key_hint = "(set)" if RESEND_API_KEY else "(empty)"
    recipient_count = len(EMAIL_TO)
    return (
        f"url={RESEND_API_URL} key={key_hint} from={EMAIL_FROM or '(empty)'} "
        f"recipients={recipient_count}"
    )


def _validate_smtp_config() -> tuple[bool, str]:
    missing = []
    if not EMAIL_ENABLED:
        missing.append("EMAIL_ENABLED")
    if not SMTP_HOST:
        missing.append("SMTP_HOST")
    if not EMAIL_FROM:
        missing.append("EMAIL_FROM")
    if not EMAIL_TO:
        missing.append("EMAIL_TO")
    if not SMTP_USERNAME:
        missing.append("SMTP_USERNAME")
    if not SMTP_PASSWORD:
        missing.append("SMTP_PASSWORD")

    if missing:
        return False, f"Missing/invalid env vars: {', '.join(missing)}"
    return True, "ok"


def _validate_resend_config() -> tuple[bool, str]:
    missing = []
    if not EMAIL_ENABLED:
        missing.append("EMAIL_ENABLED")
    if not RESEND_API_KEY:
        missing.append("RESEND_API_KEY")
    if not EMAIL_FROM:
        missing.append("EMAIL_FROM")
    if not EMAIL_TO:
        missing.append("EMAIL_TO")
    if missing:
        return False, f"Missing/invalid env vars: {', '.join(missing)}"
    return True, "ok"


def _send_email_smtp(
    subject: str,
    body: str,
    url_dates: dict[str, str] | None = None,
    *,
    html_kind: str = "auto",
) -> tuple[bool, str]:
    valid, reason = _validate_smtp_config()
    if not valid:
        logger.warning(f"Email send skipped: {reason}")
        return False, reason

    try:
        text_body, html_body = build_email_bodies(subject, body, url_dates=url_dates, html_kind=html_kind)
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = EMAIL_FROM
        msg["To"] = ", ".join(EMAIL_TO)
        msg.attach(MIMEText(text_body, "plain", _charset="utf-8"))
        msg.attach(MIMEText(html_body, "html", _charset="utf-8"))

        if SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())
        server.quit()
        return True, "email accepted by SMTP server"
    except Exception as e:
        logger.error(f"Failed to send email via SMTP: {e}")
        return False, str(e)


def _smtp_login_check() -> tuple[bool, str]:
    valid, reason = _validate_smtp_config()
    if not valid:
        return False, reason
    try:
        if SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.noop()
        server.quit()
        return True, "ok"
    except Exception as e:
        return False, str(e)


def _send_email_resend(
    subject: str,
    body: str,
    url_dates: dict[str, str] | None = None,
    *,
    html_kind: str = "auto",
) -> tuple[bool, str]:
    valid, reason = _validate_resend_config()
    if not valid:
        logger.warning(f"Email send skipped: {reason}")
        return False, reason

    try:
        text_body, html_body = build_email_bodies(subject, body, url_dates=url_dates, html_kind=html_kind)
        response = httpx.post(
            RESEND_API_URL,
            headers={
                "Authorization": f"Bearer {RESEND_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "from": EMAIL_FROM,
                "to": EMAIL_TO,
                "subject": subject,
                "text": text_body,
                "html": html_body,
            },
            timeout=30,
        )
        if response.status_code not in (200, 201, 202):
            return False, f"Resend API {response.status_code}: {response.text[:300]}"
        return True, "email accepted by Resend API"
    except Exception as e:
        logger.error(f"Failed to send email via Resend: {e}")
        return False, str(e)


def smtp_health_check() -> tuple[bool, str]:
    if EMAIL_PROVIDER == "resend":
        valid, reason = _validate_resend_config()
        if not valid:
            return False, f"{reason}. Current: {_resend_config_summary()}"
        try:
            # Check API reachability + key validity via a lightweight domains call.
            response = httpx.get(
                "https://api.resend.com/domains",
                headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
                timeout=30,
            )
            if response.status_code != 200:
                return False, f"Resend health check {response.status_code}: {response.text[:300]}"
            return True, f"Resend API reachable and key accepted. {_resend_config_summary()}"
        except Exception as e:
            return False, f"Resend health check failed: {e}. {_resend_config_summary()}"

    valid, reason = _validate_smtp_config()
    if not valid:
        return False, f"{reason}. Current: {_smtp_config_summary()}"
    ok, send_reason = _smtp_login_check()
    if ok:
        return True, f"SMTP connection/login successful. {_smtp_config_summary()}"
    return False, f"SMTP health check failed: {send_reason}. {_smtp_config_summary()}"


def send_email(
    subject: str,
    body: str,
    url_dates: dict[str, str] | None = None,
    *,
    html_kind: str = "auto",
) -> tuple[bool, str]:
    if EMAIL_PROVIDER == "resend":
        return _send_email_resend(subject, body, url_dates=url_dates, html_kind=html_kind)
    return _send_email_smtp(subject, body, url_dates=url_dates, html_kind=html_kind)


# ─── Preflight Checks ────────────────────────────────────────────────────────

def _check_resend() -> tuple[bool, str]:
    """Verify Resend API key is valid AND the sending domain is verified on this account."""
    if not EMAIL_ENABLED:
        return False, "Email is disabled (EMAIL_ENABLED not set)"
    if not RESEND_API_KEY:
        return False, "RESEND_API_KEY is not set"
    if not EMAIL_FROM:
        return False, "EMAIL_FROM is not set"
    try:
        r = httpx.get(
            "https://api.resend.com/domains",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
            timeout=15,
        )
        if r.status_code == 401:
            return False, "Resend API key is invalid or rejected (401)"
        if r.status_code != 200:
            return False, f"Resend API returned {r.status_code}: {r.text[:200]}"
        from_domain = EMAIL_FROM.split("@")[-1]
        domains = r.json().get("data", [])
        verified = any(
            d.get("name") == from_domain and d.get("status") == "verified"
            for d in domains
        )
        if not verified:
            found = [d.get("name") for d in domains] or ["none"]
            return (
                False,
                f"Domain '{from_domain}' is not verified on this Resend account.\n"
                f"Domains verified on this account: {', '.join(found)}.\n"
                f"Fix: either verify '{from_domain}' at resend.com/domains, or switch "
                f"RESEND_API_KEY to the account where '{from_domain}' is already verified.",
            )
        return True, f"Resend OK — '{from_domain}' verified, key accepted"
    except Exception as e:
        return False, f"Resend unreachable: {e}"


async def _check_kimi() -> tuple[bool, str]:
    """Verify Kimi API key and model are functional with a 1-token probe."""
    if not KIMI_API_KEY:
        return False, "KIMI_API_KEY is not set"
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.post(
                KIMI_API_URL,
                headers={"Authorization": f"Bearer {KIMI_API_KEY}", "Content-Type": "application/json"},
                json={"model": KIMI_MODEL, "messages": [{"role": "user", "content": "hi"}], "max_tokens": 1},
            )
            if r.status_code == 401:
                return False, "Kimi API key is invalid or rejected (401)"
            if r.status_code == 404:
                return False, f"Kimi model '{KIMI_MODEL}' not found (404) — update KIMI_MODEL in Railway"
            if r.status_code != 200:
                return False, f"Kimi API returned {r.status_code}: {r.text[:200]}"
            return True, f"Kimi OK — model '{KIMI_MODEL}' responding"
    except Exception as e:
        return False, f"Kimi unreachable: {e}"


async def preflight_check(check_email: bool = True) -> tuple[bool, str]:
    """Run API health checks before starting a scan. Returns (ok, message).

    check_email=True  → also validates Resend (used by /email)
    check_email=False → Kimi only (used by /scan)
    """
    kimi_ok, kimi_msg = await _check_kimi()
    resend_ok, resend_msg = (True, "skipped") if not check_email else _check_resend()

    lines = []
    all_ok = True

    if not kimi_ok:
        lines.append(f"❌ Kimi: {kimi_msg}")
        all_ok = False
    else:
        lines.append(f"✅ {kimi_msg}")

    if check_email:
        if not resend_ok:
            lines.append(f"❌ Resend: {resend_msg}")
            all_ok = False
        else:
            lines.append(f"✅ {resend_msg}")

    return all_ok, "\n".join(lines)


# ─── End Preflight ────────────────────────────────────────────────────────────

def _extract_report_field(report: str, field_name: str) -> str:
    pattern = rf"^{re.escape(field_name)}\s*:\s*(.+)$"
    m = re.search(pattern, report, flags=re.IGNORECASE | re.MULTILINE)
    return m.group(1).strip() if m else "N/A"


def _extract_section(report: str, start_marker: str, end_markers: list[str]) -> str:
    start_idx = report.find(start_marker)
    if start_idx == -1:
        return ""
    start_idx += len(start_marker)

    end_idx = len(report)
    for marker in end_markers:
        idx = report.find(marker, start_idx)
        if idx != -1:
            end_idx = min(end_idx, idx)
    return report[start_idx:end_idx].strip()


def _short_link_label(label: str, url: str) -> str:
    label = (label or "").strip()
    generic = {"source", "source url", "url", "link", "sourceurl"}
    if not label or label.lower().replace(" ", "") in generic:
        host = urlparse(url).netloc.replace("www.", "")
        label = host or "source"
    return label[:28] + "..." if len(label) > 31 else label


def _compact_email_line(raw_line: str) -> str:
    line = raw_line.strip()
    if not line:
        return ""

    line = re.sub(r"^\d+\.\s*", "", line)
    line = line.lstrip("- ").strip()
    line = line.replace("**", "").replace("`", "")

    links: list[str] = []

    def _store_link(label: str, url: str) -> str:
        safe_url = html.escape(url, quote=True)
        safe_label = html.escape(_short_link_label(label, url))
        links.append(
            f"<a href=\"{safe_url}\" style=\"font-size:13px;color:#175CD3;text-decoration:none;\">{safe_label}</a>"
        )
        return ""

    # Convert markdown links and remove them from body text.
    line = re.sub(
        r"\[([^\]]+)\]\((https?://[^\s)]+)\)",
        lambda m: _store_link(m.group(1), m.group(2)),
        line,
    )
    # Convert bare URLs into short domain links and remove from body text.
    line = re.sub(
        r"https?://[^\s)]+",
        lambda m: _store_link("", m.group(0)),
        line,
    )

    # If historical fields exist, collapse into a concise sentence.
    date_match = re.search(r"\bDate\s*:\s*([0-9]{4}-[0-9]{2}-[0-9]{2})", line, flags=re.IGNORECASE)
    product_match = re.search(r"\bProduct\s*:\s*([^:]+?)(?=\s+[A-Za-z ]+:\s|$)", line, flags=re.IGNORECASE)
    sentiment_match = re.search(r"\bSentiment Summary\s*:\s*(.+)$", line, flags=re.IGNORECASE)
    if sentiment_match:
        main_text = sentiment_match.group(1).strip()
        meta = []
        if product_match:
            meta.append(product_match.group(1).strip())
        if date_match:
            meta.append(date_match.group(1).strip())
        if meta:
            main_text = f"{main_text} ({', '.join(meta)})"
    else:
        # Keep explicit date labels in rendered bullets for faster evidence-time checks.
        line = re.sub(r"\b(Source URL|Product|Sentiment Summary)\s*:\s*", "", line, flags=re.IGNORECASE)
        main_text = " ".join(line.split())

    main_text = " ".join(main_text.split())
    if len(main_text) > 240:
        main_text = main_text[:237].rstrip() + "..."

    links_html = ""
    if links:
        links_html = " <span style=\"white-space:nowrap;\">" + " · ".join(links[:2]) + "</span>"
    if not main_text:
        return links_html
    return f"{html.escape(main_text)}{links_html}"


EMAIL_FONT_STACK = "Inter, -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif"
EMAIL_PAGE_BG = "#eef2f8"
EMAIL_CARD_BG = "#ffffff"
EMAIL_BORDER = "#d9e2f2"
EMAIL_NAVY = "#0f1f47"
EMAIL_TEXT = "#101828"
EMAIL_MUTED = "#667085"
EMAIL_ACCENT = "#175CD3"
EMAIL_CONTAINER_WIDTH = "920"
BIWEEKLY_MEMORY_PATH = Path(__file__).parent / "state" / "biweekly_memory.json"
BIWEEKLY_EXCEL_PATH = Path(__file__).parent / "state" / "biweekly_reports.xlsx"
BIWEEKLY_EXCEL_HEADERS = (
    "Scan Date",
    "e-Transfer Chatter",
    "Chatter category mix (Kimi)",
    "Market Pulse",
    "Trend vs Last Scan",
    "Full Report",
)
SOURCE_LEDGER_PATH = Path(__file__).parent / "state" / "source_ledger.xlsx"
QUARTERLY_MEMORY_PATH = Path(__file__).parent / "state" / "quarterly_memory.json"

# Scheduled quarterly market-trends runs (America/Toronto calendar).
QUARTERLY_RUN_MONTH_DAY: set[tuple[int, int]] = {(11, 1), (2, 1), (5, 1), (8, 1)}


def _compact_panel(raw: str, empty_msg: str, *, max_lines: int = 3, list_mode: bool = True) -> str:
    if not raw:
        return (
            f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;"
            f"font-size:13px;line-height:1.45;color:{EMAIL_MUTED};background:#fcfdff;'>{html.escape(empty_msg)}</div>"
        )
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    items = []
    for ln in lines[:max_lines]:
        compact = _compact_email_line(ln)
        if compact:
            items.append(compact)
    if not items:
        return (
            f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;"
            f"font-size:13px;line-height:1.45;color:{EMAIL_MUTED};background:#fcfdff;'>{html.escape(empty_msg)}</div>"
        )
    if list_mode:
        joined = "".join(f"<li style='margin:0 0 4px 0;'>{it}</li>" for it in items)
        return (
            f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;background:#fcfdff;'>"
            f"<ul style='margin:0 0 0 16px;padding:0;color:{EMAIL_TEXT};font-size:13px;line-height:1.45;'>{joined}</ul>"
            "</div>"
        )
    joined = "<br>".join(items)
    return (
        f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;"
        f"font-size:13px;line-height:1.45;color:{EMAIL_TEXT};background:#fcfdff;'>{joined}</div>"
    )


def _load_biweekly_memory() -> dict:
    try:
        if not BIWEEKLY_MEMORY_PATH.exists():
            return {}
        return json.loads(BIWEEKLY_MEMORY_PATH.read_text())
    except Exception:
        return {}


def _save_biweekly_memory(themes: dict, scan_date: str, new_urls: list[str] | None = None) -> None:
    try:
        BIWEEKLY_MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        existing = _load_biweekly_memory()
        prev_urls: list[str] = existing.get("sent_urls", [])
        merged_urls = list(dict.fromkeys(prev_urls + (new_urls or [])))[-150:]
        data = {
            "last_scan_date": scan_date,
            "etransfer_themes": themes.get("etransfer_themes", []),
            "competitor_themes": themes.get("competitor_themes", []),
            "sent_urls": merged_urls,
        }
        BIWEEKLY_MEMORY_PATH.write_text(json.dumps(data, indent=2))
    except Exception as e:
        logger.warning(f"Could not save biweekly memory: {e}")


def _load_quarterly_memory() -> dict:
    try:
        if not QUARTERLY_MEMORY_PATH.exists():
            return {}
        return json.loads(QUARTERLY_MEMORY_PATH.read_text())
    except Exception:
        return {}


def _save_quarterly_memory(*, calendar_day_iso: str) -> None:
    """Persist last successful quarterly run (Toronto calendar YYYY-MM-DD)."""
    try:
        QUARTERLY_MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        data = {"last_run_calendar_day": calendar_day_iso}
        QUARTERLY_MEMORY_PATH.write_text(json.dumps(data, indent=2))
    except Exception as e:
        logger.warning(f"Could not save quarterly memory: {e}")


def _quarterly_scan_due_today(now_local: datetime) -> tuple[bool, str]:
    if (now_local.month, now_local.day) not in QUARTERLY_RUN_MONTH_DAY:
        return False, "not a scheduled quarter-start day"
    today = now_local.date().isoformat()
    last_day = (_load_quarterly_memory().get("last_run_calendar_day") or "").strip()
    if last_day == today:
        return False, f"already completed quarterly run for {today}"
    return True, "ok"


def _extract_biweekly_themes(report: str) -> dict:
    """Extract short theme labels from biweekly report sections for memory storage."""
    etransfer_raw = _extract_section(report, "e-Transfer Chatter:", ["Market Pulse:", "Trend vs Last Scan:"])
    competitor_raw = _extract_section(report, "Market Pulse:", ["Trend vs Last Scan:"])

    def _bullets_to_themes(section_text: str) -> list[str]:
        themes = []
        for line in (section_text or "").splitlines():
            stripped = line.strip()
            if stripped.startswith("- ") and "Nothing notable" not in stripped:
                quote_only = _chatter_body_after_tag(stripped)
                if " — " in quote_only:
                    quote_only = quote_only.split(" — ")[0].strip()
                theme = quote_only[:60]
                if theme:
                    themes.append(theme)
        return themes[:6]

    return {
        "etransfer_themes": _bullets_to_themes(etransfer_raw),
        "competitor_themes": _bullets_to_themes(competitor_raw),
    }


def _maybe_migrate_biweekly_excel_headers(ws) -> list:
    """Ensure ``Chatter category mix (Kimi)`` column exists; migrate legacy 5-column sheets."""
    if ws.max_row < 1:
        return list(BIWEEKLY_EXCEL_HEADERS)
    hdr = [c.value for c in ws[1]]
    if not hdr:
        return list(BIWEEKLY_EXCEL_HEADERS)
    if "Chatter category mix (Kimi)" in hdr:
        return hdr
    if len(hdr) == 5 and hdr[0] == "Scan Date" and hdr[2] == "Market Pulse":
        ws.insert_cols(3, amount=1)
        ws.cell(row=1, column=3, value="Chatter category mix (Kimi)")
        return [c.value for c in ws[1]]
    return hdr


def _append_biweekly_excel(scan_date: str, report: str) -> None:
    """Append biweekly report sections to Excel file for human review."""
    try:
        from openpyxl import Workbook, load_workbook

        etransfer_raw = _extract_section(report, "e-Transfer Chatter:", ["Market Pulse:", "Trend vs Last Scan:"])
        competitor_raw = _extract_section(report, "Market Pulse:", ["Trend vs Last Scan:"])
        trend_raw = _extract_section(report, "Trend vs Last Scan:", [])

        BIWEEKLY_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

        if BIWEEKLY_EXCEL_PATH.exists():
            wb = load_workbook(BIWEEKLY_EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Biweekly Reports"
            ws.append(list(BIWEEKLY_EXCEL_HEADERS))

        hdr = _maybe_migrate_biweekly_excel_headers(ws)
        chatter_bullets: list[str] = []
        for line in (etransfer_raw or "").splitlines():
            s = line.strip()
            if not (s.startswith("- ") or s.startswith("• ")):
                continue
            if "Nothing notable" in s:
                continue
            if not s[2:].strip():
                continue
            chatter_bullets.append(s)
        counts = _chatter_category_counts(chatter_bullets)
        total = sum(counts[k] for k in CHATTER_CAT_ORDER)
        mix_parts: list[str] = []
        if total:
            for k in CHATTER_CAT_ORDER:
                n = counts[k]
                if not n:
                    continue
                pct = round(100.0 * n / total, 1)
                mix_parts.append(f"{CHATTER_CAT_LABELS[k]} {n} ({pct}%)")
        mix_cell = "; ".join(mix_parts)

        row_map = {
            "Scan Date": scan_date,
            "e-Transfer Chatter": (etransfer_raw or "").strip(),
            "Chatter category mix (Kimi)": mix_cell,
            "Market Pulse": (competitor_raw or "").strip(),
            "Trend vs Last Scan": (trend_raw or "").strip(),
            "Full Report": report.strip(),
        }
        ws.append([row_map.get(h, "") for h in hdr])
        wb.save(BIWEEKLY_EXCEL_PATH)
        logger.info(f"Appended biweekly report to {BIWEEKLY_EXCEL_PATH}")
    except Exception as e:
        logger.warning(f"Could not append to Excel: {e}")


def _styled_raw_report_html(subject: str, body: str) -> str:
    escaped = html.escape(body)
    return f"""
<html>
  <body style="margin:0;padding:0;background:{EMAIL_PAGE_BG};font-family:{EMAIL_FONT_STACK};color:{EMAIL_TEXT};">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="padding:24px 0;">
      <tr><td align="center">
        <table role="presentation" width="{EMAIL_CONTAINER_WIDTH}" cellspacing="0" cellpadding="0" style="background:{EMAIL_CARD_BG};border-radius:14px;overflow:hidden;border:1px solid {EMAIL_BORDER};">
          <tr>
            <td style="background:{EMAIL_NAVY};color:#ffffff;padding:18px 24px;border-bottom:4px solid #fdb913;">
              <div style="font-size:24px;font-weight:700;letter-spacing:0.2px;">Interac Intelligence</div>
              <div style="font-size:13px;color:#d8def0;margin-top:6px;">{html.escape(subject)}</div>
            </td>
          </tr>
          <tr>
            <td style="padding:20px 24px;">
              <div style="font-size:15px;font-weight:700;margin-bottom:10px;">Report</div>
              <pre style="white-space:pre-wrap;background:#f9fbff;border:1px solid {EMAIL_BORDER};border-radius:10px;padding:14px;font-size:13px;line-height:1.5;color:{EMAIL_TEXT};">{escaped}</pre>
            </td>
          </tr>
        </table>
      </td></tr>
    </table>
  </body>
</html>
""".strip()


# e-Transfer Chatter email mix — heuristic labels (keyword rules on Kimi bullet text).
CHATTER_CAT_ORDER = ("praise", "comparison", "education", "blame", "thin")
CHATTER_CAT_LABELS = {
    "praise": "Praise",
    "comparison": "Comparison",
    "education": "Education",
    "blame": "Blame",
    "thin": "Thin mention",
}
CHATTER_CAT_COLORS = {
    "praise": "#039855",
    "comparison": "#5925DC",
    "education": "#175CD3",
    "blame": "#c4320a",
    "thin": "#667085",
}

_CHATTER_COMP_TOKENS = (
    "wise", "paypal", "wealthsimple", "koho", "revolut", "neo financial", "venmo", "zelle",
    "stripe", "square", "wire transfer", "western union", " vs ", " versus ", "switched to",
    "switched from", "instead of", "moved to", "alternative to", " compared ", "better than",
    "worse than", "crypto", "bitcoin",
)
_CHATTER_BLAME_TOKENS = (
    "fraud", "scam", "phish", "stolen", "unauthorized", "hack", "hold", "pending", "declined",
    "rejected", "stuck", "frozen", "delay", "didn't arrive", "not received", "limit", "fee",
    "charged", "complaint", "problem", "issue", "broken", "error", "hate", "terrible", "worst",
    "frustrat", "annoying", "disappoint", "blame", "won't let", "wouldn't", "blocked",
)
_CHATTER_PRAISE_TOKENS = (
    "love ", " great", "awesome", "thankful", " appreciate", "works great", "so easy",
    "never had an issue", "no issues", "convenient", "best service", "smooth ", "flawless",
    " easy ", "saved me", "quick and",
)
_CHATTER_EDU_TOKENS = (
    "how do ", "how does", "what is ", "what are ", "anyone know", "anyone else", "explain",
    "confused", "first time", "new to", "is it safe", " legit ", "why does", "why do ",
    "when will", "where can", "could someone", "help me understand",
)
_CHATTER_THIN_TOKENS = ("win $", "win a", "prize", "giveaway", "contest", "promo code", "free money")


def _classify_chatter_bullet_line(line: str) -> str:
    """Keyword fallback when a chatter bullet omits the Kimi ``[Tag]`` prefix."""
    low = line.lower()
    if "nothing notable" in low:
        return "thin"
    if any(t in low for t in _CHATTER_THIN_TOKENS):
        return "thin"
    if _LOW_INSIGHT_RE.search(line):
        return "thin"
    if len(line) < 38 and line.count(" ") < 6 and "http" not in low:
        return "thin"

    if any(t in low for t in _CHATTER_COMP_TOKENS):
        return "comparison"
    if any(t in low for t in _CHATTER_BLAME_TOKENS):
        return "blame"
    if "?" in line and any(t in low for t in _CHATTER_EDU_TOKENS):
        return "education"
    if any(t in low for t in _CHATTER_PRAISE_TOKENS) and "?" not in line:
        return "praise"
    if "?" in line:
        return "education"
    if len(line) < 90:
        return "thin"
    return "blame"


def _chatter_category_counts(lines: list[str]) -> dict[str, int]:
    counts = {k: 0 for k in CHATTER_CAT_ORDER}
    for ln in lines:
        counts[_chatter_category_from_line(ln)] += 1
    return counts


def _render_chatter_category_bars(counts: dict[str, int]) -> str:
    total = sum(counts.get(k, 0) for k in CHATTER_CAT_ORDER)
    if total == 0:
        return ""
    parts = [
        "<div style='margin-bottom:16px;padding:12px 14px;background:#f9fbff;border-radius:10px;"
        f"border:1px solid {EMAIL_BORDER};'>",
        "<div style='font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;"
        f"color:{EMAIL_MUTED};margin-bottom:4px;'>Chatter mix</div>",
        f"<div style='font-size:11px;color:{EMAIL_MUTED};margin-bottom:10px;line-height:1.4;'>"
        f"{html.escape('Share of bullets from Kimi [Tag] labels on each line; missing tags use a keyword fallback.')}"
        "</div>",
    ]
    for key in CHATTER_CAT_ORDER:
        n = counts.get(key, 0)
        pct = round(100.0 * n / total, 1) if total else 0.0
        label = CHATTER_CAT_LABELS[key]
        color = CHATTER_CAT_COLORS[key]
        w = min(100.0, max(0.0, pct))
        parts.append(
            f"<div style='margin-bottom:10px;'>"
            f"<table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='margin-bottom:2px;'>"
            f"<tr><td style='font-size:12px;font-weight:600;color:{EMAIL_TEXT};'>{html.escape(label)}</td>"
            f"<td align='right' style='font-size:12px;color:{EMAIL_MUTED};'>{pct:.1f}% ({n})</td></tr></table>"
            f"<div style='background:#e8ecf4;border-radius:6px;height:11px;overflow:hidden;'>"
            f"<div style='width:{w:.1f}%;background:{color};height:11px;border-radius:6px;'></div>"
            f"</div></div>"
        )
    parts.append("</div>")
    return "".join(parts)


def _render_chatter_column_with_mix(raw: str, url_dates: dict[str, str] | None) -> str:
    """e-Transfer Chatter HTML: category mix bars + quote cards with small category badges."""
    _empty = (
        f"<div style='padding:12px 14px;font-size:13px;color:{EMAIL_MUTED};'>"
        f"{html.escape('Nothing notable this scan.')}</div>"
    )
    if not raw or not raw.strip():
        return _empty

    bullet_lines: list[str] = []
    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped or not (stripped.startswith("- ") or stripped.startswith("• ")):
            continue
        text = stripped[2:].strip()
        if not text or "Nothing notable" in text:
            continue
        bullet_lines.append(stripped)

    if not bullet_lines:
        return _empty

    bars = _render_chatter_category_bars(_chatter_category_counts(bullet_lines))

    cards: list[str] = []
    for stripped in bullet_lines:
        cat = _chatter_category_from_line(stripped)
        cat_color = CHATTER_CAT_COLORS[cat]
        cat_label = CHATTER_CAT_LABELS[cat]
        text = _chatter_body_after_tag(stripped)

        if " — " in text:
            quote_part, attr_part = text.split(" — ", 1)
        else:
            quote_part, attr_part = text, ""

        url = ""
        url_match = re.search(r"Source:\s*(https?://\S+)", attr_part, re.IGNORECASE)
        if url_match:
            url = url_match.group(1).rstrip(".,)")
            attr_part = re.sub(r"\s*Source:\s*https?://\S+", "", attr_part, flags=re.IGNORECASE).strip()
        else:
            bare_match = re.search(r"(https?://\S+)", attr_part)
            if bare_match:
                url = bare_match.group(1).rstrip(".,)")
                attr_part = re.sub(r"https?://\S+", "", attr_part).strip()

        attr_clean = attr_part.rstrip(".")
        if "," in attr_clean:
            _platform_from_llm, date_label = attr_clean.split(",", 1)
            date_label = date_label.strip().rstrip(".")
        else:
            date_label = ""

        if not date_label and url:
            date_label = (url_dates or {}).get(_canonical_url_for_date_lookup(url), "")
        if not date_label and url:
            date_label = _resolve_relative_date(_extract_date_from_url(url))

        COMMUNITY_SOURCES = {"Reddit", "X/Twitter", "RedFlagDeals", "Forum"}
        show_badge = False
        platform_label = ""
        if url:
            _, url_source = _classify_channel_and_source(url)
            if url_source in COMMUNITY_SOURCES:
                platform_label = url_source
                show_badge = True

        badge_color = _platform_badge_color(platform_label) if show_badge else "#d1d5db"

        link_html = ""
        if url:
            safe_url = html.escape(url, quote=True)
            domain = re.sub(r"^www\.", "", re.sub(r"https?://", "", url).split("/")[0])
            link_html = (
                f"<a href='{safe_url}' style='font-size:11px;color:{EMAIL_ACCENT};"
                f"text-decoration:none;'>{html.escape(domain)}</a>"
            )

        quote_html = html.escape(quote_part.strip())
        date_html = html.escape(date_label) if date_label else ""

        cat_badge = (
            f"<span style='font-size:9px;font-weight:700;padding:2px 7px;border-radius:4px;"
            f"letter-spacing:0.2px;background:{cat_color}18;color:{cat_color};"
            f"margin-bottom:6px;display:inline-block;'>{html.escape(cat_label)}</span>"
        )

        meta_inner = ""
        if show_badge and platform_label:
            meta_inner += (
                f"<span style='background:{badge_color};color:#fff;font-size:10px;"
                f"font-weight:700;padding:2px 7px;border-radius:999px;"
                f"letter-spacing:0.3px;white-space:nowrap;margin-right:6px;display:inline-block;'>"
                f"{html.escape(platform_label)}</span>"
            )
        if date_html:
            meta_inner += (
                f"<span style='font-size:11px;color:{EMAIL_MUTED};margin-right:6px;'>{date_html}</span>"
            )
        if link_html:
            meta_inner += link_html

        meta_row = (
            f"<div style='margin-top:5px;line-height:1.8;'>{meta_inner}</div>" if meta_inner else ""
        )

        card = (
            f"<div style='border-left:3px solid {cat_color};padding:8px 0 8px 12px;margin-bottom:14px;'>"
            f"{cat_badge}"
            f"<div style='font-size:13px;line-height:1.55;color:{EMAIL_TEXT};'>{quote_html}</div>"
            f"{meta_row}"
            f"</div>"
        )
        cards.append(card)

    return bars + "".join(cards)


def _platform_badge_color(platform: str) -> str:
    """Return badge color for a platform name."""
    p = platform.lower()
    if "reddit" in p:
        return "#FF4500"
    if "redflagdeal" in p or "rfd" in p:
        return "#CC1200"
    if "twitter" in p or "x/" in p or p == "x":
        return "#1a1a1a"
    if "forum" in p:
        return "#6b7280"
    return "#1a73e8"  # News/Other


def _render_quote_bullets(
    raw: str, empty_msg: str, url_dates: dict[str, str] | None = None
) -> str:
    """Render quote bullets as styled cards with platform badge, date, and hyperlink.

    Each bullet is expected in the form:
        - "quote text" — Platform, Date. Source: URL
    When the model omits the date, ``url_dates`` (canonical URL -> date) can fill it in.
    """
    _empty = (
        f"<div style='padding:12px 14px;font-size:13px;color:{EMAIL_MUTED};'>"
        f"{html.escape(empty_msg)}</div>"
    )
    if not raw:
        return _empty

    items = []
    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped or not (stripped.startswith("- ") or stripped.startswith("• ")):
            continue
        text = stripped[2:].strip()
        if not text or "Nothing notable" in text:
            continue

        # Split quote from attribution on ' — '
        if " — " in text:
            quote_part, attr_part = text.split(" — ", 1)
        else:
            quote_part, attr_part = text, ""

        # Extract URL — try "Source: URL" first, then bare https?:// in attr
        url = ""
        url_match = re.search(r"Source:\s*(https?://\S+)", attr_part, re.IGNORECASE)
        if url_match:
            url = url_match.group(1).rstrip(".,)")
            attr_part = re.sub(r"\s*Source:\s*https?://\S+", "", attr_part, flags=re.IGNORECASE).strip()
        else:
            bare_match = re.search(r"(https?://\S+)", attr_part)
            if bare_match:
                url = bare_match.group(1).rstrip(".,)")
                attr_part = re.sub(r"https?://\S+", "", attr_part).strip()

        # Parse date from attribution (everything after first comma).
        # Format with date:    — Platform, Date.
        # Format without date: — Platform.  (no comma, so no date)
        attr_clean = attr_part.rstrip(".")
        if "," in attr_clean:
            _platform_from_llm, date_label = attr_clean.split(",", 1)
            date_label = date_label.strip().rstrip(".")
        else:
            # No comma → attribution is platform only, no date
            date_label = ""

        if not date_label and url:
            date_label = (url_dates or {}).get(_canonical_url_for_date_lookup(url), "")
        if not date_label and url:
            date_label = _resolve_relative_date(_extract_date_from_url(url))

        # Derive platform badge from URL domain — only community platforms get a badge.
        # This prevents corporate sites (wise.com, paypal.com) from showing as badges.
        COMMUNITY_SOURCES = {"Reddit", "X/Twitter", "RedFlagDeals", "Forum"}
        show_badge = False
        platform_label = ""
        if url:
            _, url_source = _classify_channel_and_source(url)
            if url_source in COMMUNITY_SOURCES:
                platform_label = url_source
                show_badge = True

        badge_color = _platform_badge_color(platform_label) if show_badge else "#d1d5db"

        # Build link HTML
        link_html = ""
        if url:
            safe_url = html.escape(url, quote=True)
            domain = re.sub(r"^www\.", "", re.sub(r"https?://", "", url).split("/")[0])
            link_html = (
                f"<a href='{safe_url}' style='font-size:11px;color:{EMAIL_ACCENT};"
                f"text-decoration:none;'>{html.escape(domain)}</a>"
            )

        quote_html = html.escape(quote_part.strip())
        date_html = html.escape(date_label) if date_label else ""

        # Build meta row — margin-right on each element (flex gap unreliable in webmail)
        meta_inner = ""
        if show_badge and platform_label:
            meta_inner += (
                f"<span style='background:{badge_color};color:#fff;font-size:10px;"
                f"font-weight:700;padding:2px 7px;border-radius:999px;"
                f"letter-spacing:0.3px;white-space:nowrap;margin-right:6px;display:inline-block;'>"
                f"{html.escape(platform_label)}</span>"
            )
        if date_html:
            meta_inner += (
                f"<span style='font-size:11px;color:{EMAIL_MUTED};margin-right:6px;'>{date_html}</span>"
            )
        if link_html:
            meta_inner += link_html

        meta_row = (
            f"<div style='margin-top:5px;line-height:1.8;'>{meta_inner}</div>"
            if meta_inner else ""
        )

        card = (
            f"<div style='border-left:3px solid {badge_color};padding:8px 0 8px 12px;"
            f"margin-bottom:14px;'>"
            f"<div style='font-size:13px;line-height:1.55;color:{EMAIL_TEXT};'>{quote_html}</div>"
            f"{meta_row}"
            f"</div>"
        )
        items.append(card)

    if not items:
        return _empty

    return "".join(items)


def _trend_mini_card(label: str, content: str, accent: str) -> str:
    """Render a single Trend sub-column card."""
    safe_content = html.escape(content.strip()) if content.strip() else "none identified"
    return (
        f"<td style='width:33%;vertical-align:top;padding:0 8px 0 0;'>"
        f"<div style='background:#eef2f8;border-radius:10px;padding:12px 14px;height:100%;box-sizing:border-box;'>"
        f"<div style='font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;"
        f"color:{accent};margin-bottom:6px;'>{html.escape(label)}</div>"
        f"<div style='font-size:12px;line-height:1.55;color:{EMAIL_TEXT};'>{safe_content}</div>"
        f"</div></td>"
    )


def _parse_trend_fields(trend_raw: str) -> tuple[str, str, str]:
    """Extract Still active / Went quiet / New this scan values from trend section."""
    still, quiet, new = "", "", ""
    for line in (trend_raw or "").splitlines():
        l = line.strip()
        if l.lower().startswith("- still active:"):
            still = l.split(":", 1)[1].strip()
        elif l.lower().startswith("- went quiet:"):
            quiet = l.split(":", 1)[1].strip()
        elif l.lower().startswith("- new this scan:"):
            new = l.split(":", 1)[1].strip()
    return still, quiet, new


def _build_biweekly_html(
    subject: str, body: str, url_dates: dict[str, str] | None = None
) -> str:
    scan_date = _extract_report_field(body, "SCAN DATE")
    etransfer_raw = _extract_section(body, "e-Transfer Chatter:", ["Market Pulse:", "Trend vs Last Scan:"])
    competitor_raw = _extract_section(body, "Market Pulse:", ["Trend vs Last Scan:"])
    if not any(s.strip() for s in [etransfer_raw, competitor_raw]):
        return _styled_raw_report_html(subject, body)

    umap = url_dates or {}
    etransfer_html = _render_chatter_column_with_mix(etransfer_raw, umap)
    competitor_html = _render_quote_bullets(competitor_raw, "Nothing notable this scan.", umap)

    trend_block = _extract_section(body, "Trend vs Last Scan:", [])
    ledger_line = f"Source ledger: {source_ledger_display_url()}"
    if trend_block:
        tl = trend_block.lower()
        if "source ledger:" in tl:
            idx = tl.rfind("source ledger:")
            ledger_line = trend_block[idx:].strip()
            trend_block = trend_block[:idx].strip()
        still, quiet, new = _parse_trend_fields(trend_block)
        trend_html = (
            f"<tr><td colspan='2' style='padding:18px 28px 8px 28px;border-top:1px solid {EMAIL_BORDER};'>"
            f"<div style='font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;"
            f"color:{EMAIL_MUTED};margin-bottom:10px;'>Trend vs last scan</div>"
            f"<table role='presentation' width='100%' cellspacing='0' cellpadding='0'><tr>"
            f"{_trend_mini_card('Still active', still, '#175CD3')}"
            f"{_trend_mini_card('Went quiet', quiet, '#667085')}"
            f"{_trend_mini_card('New this scan', new, '#039855')}"
            f"</tr></table></td></tr>"
        )
    else:
        trend_html = ""

    ledger_html = (
        f"<tr><td colspan='2' style='padding:12px 28px 22px 28px;border-top:1px solid {EMAIL_BORDER};"
        f"background:#f9fbff;'>"
        f"<div style='font-size:13px;line-height:1.5;color:{EMAIL_TEXT};'>{html.escape(ledger_line)}</div>"
        f"</td></tr>"
    )

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  .qcard:hover {{ background:#f5f8ff !important; }}
  a:hover {{ text-decoration:underline !important; }}
</style>
</head>
<body style="margin:0;padding:0;background:{EMAIL_PAGE_BG};font-family:{EMAIL_FONT_STACK};color:{EMAIL_TEXT};">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="padding:24px 0;">
    <tr><td align="center">
      <table role="presentation" width="1200" cellspacing="0" cellpadding="0" style="background:{EMAIL_CARD_BG};border-radius:14px;overflow:hidden;border:1px solid {EMAIL_BORDER};">

        <!-- HEADER -->
        <tr>
          <td colspan="2" style="background:{EMAIL_NAVY};color:#ffffff;padding:20px 28px;border-bottom:4px solid #fdb913;">
            <div style="font-size:24px;font-weight:700;letter-spacing:0.2px;">Interac e-Transfer Intelligence</div>
            <div style="font-size:13px;color:#aebce2;margin-top:5px;">{html.escape(scan_date)}</div>
          </td>
        </tr>

        <!-- TWO-COLUMN BODY -->
        <tr>
          <!-- LEFT: e-Transfer Chatter -->
          <td width="50%" style="vertical-align:top;padding:22px 14px 22px 28px;border-right:1px solid {EMAIL_BORDER};">
            <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;color:#c4320a;margin-bottom:4px;">Pain points</div>
            <div style="font-size:16px;font-weight:700;color:{EMAIL_TEXT};margin-bottom:16px;">e-Transfer Chatter</div>
            {etransfer_html}
          </td>
          <!-- RIGHT: Market Pulse -->
          <td width="50%" style="vertical-align:top;padding:22px 28px 22px 14px;">
            <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;color:#5925DC;margin-bottom:4px;">Market pulse</div>
            <div style="font-size:16px;font-weight:700;color:{EMAIL_TEXT};margin-bottom:16px;">Payments Landscape</div>
            {competitor_html}
          </td>
        </tr>
        {trend_html}
        {ledger_html}

      </table>
    </td></tr>
  </table>
</body>
</html>""".strip()


def _build_quarterly_html(subject: str, body: str, url_dates: dict[str, str] | None = None) -> str:
    """Sectioned HTML for long-form quarterly reports (split on markdown ### headings)."""
    scan_date = _extract_report_field(body, "REPORT DATE") or _extract_report_field(body, "SCAN DATE")
    raw = body.strip()
    parts = re.split(r"(?m)^###\s+", raw)
    blocks: list[tuple[str, str]] = []
    if parts and parts[0].strip():
        blocks.append(("Overview", parts[0].strip()))
    for chunk in parts[1:]:
        if not chunk.strip():
            continue
        first_line, _, rest = chunk.partition("\n")
        title = first_line.strip() or "Section"
        blocks.append((title, rest.strip() if rest.strip() else chunk.strip()))

    sections_html = ""
    for title, content in blocks:
        sections_html += (
            f"<div style='margin-bottom:18px;border:1px solid {EMAIL_BORDER};border-radius:10px;"
            f"padding:14px 16px;background:#fcfdff;'>"
            f"<h3 style='margin:0 0 8px 0;font-size:15px;color:{EMAIL_NAVY};'>{html.escape(title)}</h3>"
            f"<pre style='white-space:pre-wrap;font-size:13px;line-height:1.5;margin:0;color:{EMAIL_TEXT};'>"
            f"{html.escape(content)}</pre></div>"
        )

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
</head>
<body style="margin:0;padding:0;background:{EMAIL_PAGE_BG};font-family:{EMAIL_FONT_STACK};color:{EMAIL_TEXT};">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="padding:24px 0;">
    <tr><td align="center">
      <table role="presentation" width="{EMAIL_CONTAINER_WIDTH}" cellspacing="0" cellpadding="0" style="background:{EMAIL_CARD_BG};border-radius:14px;overflow:hidden;border:1px solid {EMAIL_BORDER};">
        <tr>
          <td style="background:{EMAIL_NAVY};color:#ffffff;padding:20px 28px;border-bottom:4px solid #fdb913;">
            <div style="font-size:24px;font-weight:700;letter-spacing:0.2px;">Quarterly market trends</div>
            <div style="font-size:13px;color:#aebce2;margin-top:5px;">{html.escape(scan_date)}</div>
            <div style="font-size:12px;color:#d8def0;margin-top:6px;">{html.escape(subject)}</div>
          </td>
        </tr>
        <tr>
          <td style="padding:22px 28px;vertical-align:top;">
            {sections_html}
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>""".strip()


def build_email_bodies(
    subject: str,
    body: str,
    url_dates: dict[str, str] | None = None,
    *,
    html_kind: str = "auto",
) -> tuple[str, str]:
    kind = html_kind
    if kind == "auto":
        kind = (
            "quarterly"
            if "QUARTERLY" in subject.upper() or "Quarterly market trends" in subject
            else "biweekly"
        )
    if kind == "quarterly":
        merged = url_dates if url_dates is not None else last_quarterly_url_dates
        return body, _build_quarterly_html(subject, body, url_dates=merged or {})
    merged = url_dates if url_dates is not None else last_biweekly_url_dates
    return body, _build_biweekly_html(subject, body, url_dates=merged)


def _record_email_sent(trigger: str, *, now_local: datetime | None = None) -> None:
    global last_email_sent_at, last_weekly_email_key, last_quarterly_email_key
    last_email_sent_at = datetime.now(timezone.utc)
    if trigger == "weekly" and now_local is not None:
        last_weekly_email_key = weekly_key(now_local)
    if trigger == "quarterly" and now_local is not None:
        last_quarterly_email_key = f"q-{now_local.date().isoformat()}"


def weekly_est_to_utc(day_name: str, hour_est: int) -> tuple[int, int]:
    base_day = WEEKDAY_TO_INDEX.get(day_name, 0)
    hour_utc = hour_est + 5  # EST -> UTC
    day_shift = 0
    if hour_utc >= 24:
        hour_utc -= 24
        day_shift = 1
    return (base_day + day_shift) % 7, hour_utc


async def ask_followup(question: str, report_context: str) -> str:
    config = load_prompts()
    return await call_kimi(
        config["followup_prompt"],
        f"Latest report:\n{report_context}\n\nRaw mentions:\n{last_mentions_raw[:3000]}\n\nQuestion: {question}",
    )


# Telegram hard-limits message length (4096 chars). Chunk long reports safely.
async def send_chunked_message(
    update: Update,
    text: str,
    *,
    parse_mode: str | None = None,
    chunk_size: int = 3900,
) -> None:
    async def _reply_with_fallback(message_text: str) -> None:
        if not parse_mode:
            await update.message.reply_text(message_text)
            return
        try:
            await update.message.reply_text(message_text, parse_mode=parse_mode)
        except BadRequest as e:
            # Model-generated text can contain invalid markdown entities.
            # Retry as plain text so delivery succeeds instead of failing the command.
            if "Can't parse entities" in str(e):
                await update.message.reply_text(message_text)
                return
            raise

    if len(text) <= chunk_size:
        await _reply_with_fallback(text)
        return

    chunks = []
    remaining = text
    while remaining:
        if len(remaining) <= chunk_size:
            chunks.append(remaining)
            break
        split_at = remaining.rfind("\n\n", 0, chunk_size)
        if split_at < 0:
            split_at = remaining.rfind("\n", 0, chunk_size)
        if split_at < 0:
            split_at = chunk_size
        chunks.append(remaining[:split_at].rstrip())
        remaining = remaining[split_at:].lstrip()

    total = len(chunks)
    for idx, chunk in enumerate(chunks, 1):
        prefix = f"({idx}/{total})\n" if total > 1 else ""
        await _reply_with_fallback(prefix + chunk)


async def send_chunked_plain_chat(
    context: ContextTypes.DEFAULT_TYPE, chat_id: int, text: str, chunk_size: int = 3900
) -> None:
    """Send a long plain-text message to a chat (e.g. scheduled jobs), respecting Telegram limits."""
    if len(text) <= chunk_size:
        await context.bot.send_message(chat_id=chat_id, text=text)
        return
    chunks: list[str] = []
    remaining = text
    while remaining:
        if len(remaining) <= chunk_size:
            chunks.append(remaining)
            break
        split_at = remaining.rfind("\n\n", 0, chunk_size)
        if split_at < 0:
            split_at = remaining.rfind("\n", 0, chunk_size)
        if split_at < 0:
            split_at = chunk_size
        chunks.append(remaining[:split_at].rstrip())
        remaining = remaining[split_at:].lstrip()
    total = len(chunks)
    for idx, chunk in enumerate(chunks, 1):
        prefix = f"({idx}/{total})\n" if total > 1 else ""
        await context.bot.send_message(chat_id=chat_id, text=prefix + chunk)


# ─── Scheduled Broadcast ─────────────────────────────────────────────────────
async def scheduled_biweekly_broadcast(context: ContextTypes.DEFAULT_TYPE):
    """Daily job that runs the biweekly scan if 14+ days have passed since the last one."""
    global last_report, last_mentions_raw

    memory = _load_biweekly_memory()
    last_date_str = memory.get("last_scan_date")
    if last_date_str:
        try:
            last_date = datetime.fromisoformat(last_date_str)
            if not last_date.tzinfo:
                last_date = last_date.replace(tzinfo=EST)
            days_since = (datetime.now(EST) - last_date).days
            if days_since < 14:
                logger.info(f"Biweekly scan skipped: {days_since} days since last scan (need 14).")
                return
        except Exception as e:
            logger.warning(f"Could not parse last_scan_date for biweekly guard: {e}")

    tracked = _track_current_task()
    logger.info(f"[{now_est()}] Running scheduled biweekly scan...")
    try:
        mentions, sources = await fetch_biweekly_mentions()
        last_mentions_raw = mentions
        report = await analyze_biweekly(mentions, sources)
        last_report = report

        message = f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}"
        for chat_id in subscribed_chats.copy():
            try:
                await context.bot.send_message(chat_id=chat_id, text=message)
            except Exception as e:
                logger.error(f"Failed to send biweekly report to {chat_id}: {e}")
                subscribed_chats.discard(chat_id)

        now_local = datetime.now(EST)
        should_send, reason = _should_send_email(trigger="weekly", now_local=now_local)
        if should_send:
            subject = f"{EMAIL_SUBJECT_PREFIX} — BIWEEKLY REPORT"
            ok, send_reason = send_email(
                subject,
                f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}",
                html_kind="biweekly",
            )
            if ok:
                _record_email_sent("weekly", now_local=now_local)
            else:
                logger.error(f"Biweekly email failed: {send_reason}")
        else:
            logger.info(f"Biweekly email not sent: {reason}")
    except Exception as e:
        logger.error(f"Scheduled biweekly scan failed: {e}")
    finally:
        _untrack_task(tracked)


async def scheduled_quarterly_market_trends(context: ContextTypes.DEFAULT_TYPE):
    """Daily job: on Nov 1 / Feb 1 / May 1 / Aug 1 (Toronto), run the quarterly trends report once."""
    now_local = datetime.now(EST)
    due, reason = _quarterly_scan_due_today(now_local)
    if not due:
        logger.info(f"Quarterly market trends skipped: {reason}")
        return

    tracked = _track_current_task()
    logger.info(f"[{now_est()}] Running scheduled quarterly market trends scan...")
    try:
        mentions, sources = await asyncio.wait_for(fetch_biweekly_mentions(quarterly=True), timeout=600)
        if mentions.startswith("No mentions"):
            logger.warning(f"Quarterly scan: no data — {mentions[:200]}")
            _save_quarterly_memory(calendar_day_iso=now_local.date().isoformat())
            return

        report, _digest, _used = await asyncio.wait_for(analyze_quarterly(mentions, sources), timeout=600)

        message = f"Interac e-Transfer — Quarterly market trends — {now_est()}\n\n{report}"
        for chat_id in subscribed_chats.copy():
            try:
                await send_chunked_plain_chat(context, chat_id, message)
            except Exception as e:
                logger.error(f"Failed to send quarterly report to {chat_id}: {e}")
                subscribed_chats.discard(chat_id)

        should_send, email_reason = _should_send_email(trigger="quarterly", now_local=now_local)
        if should_send:
            subject = f"{EMAIL_SUBJECT_PREFIX} — QUARTERLY MARKET TRENDS"
            ok, send_reason = send_email(
                subject,
                message,
                url_dates=last_quarterly_url_dates,
                html_kind="quarterly",
            )
            if ok:
                _record_email_sent("quarterly", now_local=now_local)
            else:
                logger.error(f"Quarterly email failed: {send_reason}")
        else:
            logger.info(f"Quarterly email not sent: {email_reason}")

        _save_quarterly_memory(calendar_day_iso=now_local.date().isoformat())
    except asyncio.TimeoutError:
        logger.error("Quarterly market trends scan timed out.")
    except Exception as e:
        logger.error(f"Scheduled quarterly market trends failed: {e}")
    finally:
        _untrack_task(tracked)


# ─── Command Handlers ────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribed_chats.add(update.effective_chat.id)
    await update.message.reply_text(
        "👋 *Interac Intelligence Bot*\n\n"
        "Scans Reddit, X, RedFlagDeals, and news for e-Transfer chatter. "
        "Biweekly report delivered automatically.\n\n"
        "*Commands:*\n"
        "• /subscribe — Get scheduled biweekly reports\n"
        "• /unsubscribe — Stop reports\n"
        "• /scan — Run biweekly scan now\n"
        "• /raw — See raw mentions from last scan\n"
        "• /prompt — View current config\n"
        "• /status — Check bot status\n"
        "• /email — Admin: run fresh biweekly scan + send email\n"
        "• /quarterly — Admin: quarterly market trends report\n"
        "• /stop — Admin: cancel running jobs\n"
        "• /smtpcheck — Admin: check email config\n"
        "• Any text → Follow-up on latest report",
        parse_mode="Markdown",
    )


async def cmd_subscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribed_chats.add(update.effective_chat.id)
    await update.message.reply_text("✅ Subscribed.")


async def cmd_unsubscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribed_chats.discard(update.effective_chat.id)
    await update.message.reply_text("🔕 Unsubscribed.")


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    config = load_prompts()
    et_count = len(config.get("etransfer_queries", config.get("biweekly_queries", [])))
    comp_count = len(config.get("competitor_queries", []))
    memory = _load_biweekly_memory()
    qmem = _load_quarterly_memory()
    last_scan = memory.get("last_scan_date", "never")
    last_q = qmem.get("last_run_calendar_day", "never")
    is_admin = "✅" if update.effective_user.id in ADMIN_IDS else "❌"
    resend_key_hint = f"✅ set ({RESEND_API_KEY[:6]}...)" if RESEND_API_KEY else "❌ missing"
    kimi_key_hint = f"✅ set ({KIMI_API_KEY[:6]}...)" if KIMI_API_KEY else "❌ missing"
    twitter_key_hint = f"✅ twitterapi.io ({TWITTERAPI_IO_KEY[:6]}...)" if TWITTERAPI_IO_KEY else "⚠️ DDG fallback"
    await update.message.reply_text(
        f"✅ Bot running — {now_est()}\n"
        f"Search provider: DuckDuckGo (ddgs)\n"
        f"e-Transfer queries: {et_count} | Competitor queries: {comp_count}\n"
        f"Last biweekly scan: {last_scan}\n"
        f"Last quarterly market trends run (calendar day): {last_q}\n"
        f"Schedule: daily check at 9am EST, runs every 14 days\n"
        f"Quarterly: Nov 1 / Feb 1 / May 1 / Aug 1 (same daily job time)\n"
        f"Subscribed chats: {len(subscribed_chats)}\n"
        f"Email provider: {EMAIL_PROVIDER}\n"
        f"Resend key: {resend_key_hint}\n"
        f"Kimi key: {kimi_key_hint}\n"
        f"Twitter: {twitter_key_hint}\n"
        f"Email from: {EMAIL_FROM or '❌ missing'}\n"
        f"Email to: {EMAIL_TO or '❌ missing'}\n"
        f"Admin: {is_admin}\n"
        f"Your ID: `{update.effective_user.id}`"
    )


async def cmd_scan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ok, msg = await preflight_check(check_email=False)
    if not ok:
        await update.message.reply_text(f"❌ Preflight failed — scan aborted:\n\n{msg}")
        return
    try:
        await run_biweekly_scan(update)
    except asyncio.TimeoutError:
        await update.message.reply_text("⏱️ /scan timed out after 480 seconds.")
    except Exception as e:
        logger.error(f"Scan failed: {e}")
        await update.message.reply_text(f"❌ Scan failed: {e}")


async def cmd_raw(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not last_mentions_raw:
        await update.message.reply_text("No scan data yet. Run /scan first.")
        return
    text = last_mentions_raw[:4000]
    await update.message.reply_text(f"```\n{text}\n```", parse_mode="Markdown")


async def cmd_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    config = load_prompts()
    et_queries = config.get("etransfer_queries", config.get("biweekly_queries", []))
    comp_queries = config.get("competitor_queries", [])
    sample_et = "\n".join(f"  • {q}" for q in et_queries[:3])
    sample_comp = "\n".join(f"  • {q}" for q in comp_queries[:3])
    await update.message.reply_text(
        f"*e-Transfer queries:* {len(et_queries)}\n{sample_et}\n\n"
        f"*Competitor queries:* {len(comp_queries)}\n{sample_comp}\n\n"
        f"Edit `prompts.json` to change queries.\n"
        f"Edit `prompts/biweekly_prompt.md` to change the report format.",
        parse_mode="Markdown",
    )


async def cmd_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global last_report, last_mentions_raw
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    await update.message.reply_text("🔍 Running preflight checks...")
    ok, msg = await preflight_check(check_email=True)
    if not ok:
        await update.message.reply_text(f"❌ Preflight failed — scan aborted:\n\n{msg}")
        return
    await update.message.reply_text(f"✅ Preflight passed:\n{msg}\n\nStarting scan...")

    tracked = _track_current_task()
    await update.message.reply_text("📧 Running fresh biweekly scan and sending email...")
    try:
        mentions, sources = await asyncio.wait_for(fetch_biweekly_mentions(), timeout=300)
        last_mentions_raw = mentions

        if mentions.startswith("No mentions"):
            await update.message.reply_text(f"No data found this scan.\n\n{mentions[:500]}")
            return

        await update.message.reply_text("Mentions collected. Running Kimi curation pass, then analysis...")
        report = await asyncio.wait_for(analyze_biweekly(mentions, sources), timeout=180)
        last_report = report

        subject = f"{EMAIL_SUBJECT_PREFIX} — MANUAL REPORT"
        body = f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}"
        ok, send_reason = send_email(subject, body, html_kind="biweekly")
        if ok:
            _record_email_sent("on_demand")
            await update.message.reply_text("✅ Email sent successfully.")
        else:
            await update.message.reply_text(f"❌ Email failed: {send_reason}")
    except asyncio.TimeoutError:
        await update.message.reply_text("⏱️ /email timed out after 480 seconds.")
    except Exception as e:
        logger.error(f"/email failed: {e}")
        await update.message.reply_text(f"❌ /email failed: {e}")
    finally:
        _untrack_task(tracked)


async def cmd_quarterly(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin: run the quarterly market-trends fetch + Kimi report (long timeout)."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    ok, msg = await preflight_check(check_email=False)
    if not ok:
        await update.message.reply_text(f"❌ Preflight failed — aborted:\n\n{msg}")
        return

    await update.message.reply_text(
        "Running quarterly market trends scan (~90-day pool). This may take several minutes..."
    )
    tracked = _track_current_task()
    try:
        mentions, sources = await asyncio.wait_for(fetch_biweekly_mentions(quarterly=True), timeout=600)
        if mentions.startswith("No mentions"):
            await update.message.reply_text(f"No data found.\n\n{mentions[:800]}")
            return

        await update.message.reply_text("Mentions collected. Running quarterly Kimi analysis...")
        report, _digest, _used = await asyncio.wait_for(analyze_quarterly(mentions, sources), timeout=600)

        out = f"Interac e-Transfer — Quarterly market trends — {now_est()}\n\n{report}"
        await send_chunked_message(update, out)

        now_local = datetime.now(EST)
        should_send, email_reason = _should_send_email(trigger="quarterly", now_local=now_local)
        if should_send:
            subject = f"{EMAIL_SUBJECT_PREFIX} — QUARTERLY MARKET TRENDS (manual)"
            ok_mail, send_reason = send_email(
                subject,
                out,
                url_dates=last_quarterly_url_dates,
                html_kind="quarterly",
            )
            if ok_mail:
                _record_email_sent("quarterly", now_local=now_local)
                await update.message.reply_text("✅ Quarterly email sent.")
            else:
                await update.message.reply_text(f"⚠️ Quarterly email not sent: {send_reason}")
        else:
            await update.message.reply_text(f"ℹ️ Email skipped: {email_reason}")

        _save_quarterly_memory(calendar_day_iso=now_local.date().isoformat())
    except asyncio.TimeoutError:
        await update.message.reply_text("⏱️ /quarterly timed out after 20 minutes.")
    except Exception as e:
        logger.error(f"/quarterly failed: {e}")
        await update.message.reply_text(f"❌ /quarterly failed: {e}")
    finally:
        _untrack_task(tracked)


async def cmd_stop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    current = asyncio.current_task()
    cancelled = _cancel_active_tasks(exclude=current)
    await update.message.reply_text(f"🛑 Stop requested. Cancelled {cancelled} running task(s).")


async def cmd_smtpcheck(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    ok, reason = smtp_health_check()
    if ok:
        await update.message.reply_text(f"✅ {reason}")
    else:
        await update.message.reply_text(f"❌ {reason}")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_text = update.message.text
    if not user_text:
        return
    if not last_report:
        await update.message.reply_text("No report yet. Run /scan first.")
        return

    allowed, remaining = check_rate_limit(update.effective_user.id)
    if not allowed:
        await update.message.reply_text(
            f"⚠️ Daily limit reached ({DAILY_LIMIT} questions/day). Resets at midnight EST."
        )
        return

    await update.message.reply_text("🤔 Thinking...")
    try:
        response = await ask_followup(user_text, last_report)
        suffix = f"\n\n_({remaining} questions remaining today)_" if remaining >= 0 else ""
        await update.message.reply_text(response + suffix, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_start))
    app.add_handler(CommandHandler("subscribe", cmd_subscribe))
    app.add_handler(CommandHandler("unsubscribe", cmd_unsubscribe))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("scan", cmd_scan))
    app.add_handler(CommandHandler("raw", cmd_raw))
    app.add_handler(CommandHandler("prompt", cmd_prompt))
    app.add_handler(CommandHandler("email", cmd_email))
    app.add_handler(CommandHandler("quarterly", cmd_quarterly))
    app.add_handler(CommandHandler("stop", cmd_stop))
    app.add_handler(CommandHandler("smtpcheck", cmd_smtpcheck))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # Biweekly scan: runs daily at 9am EST (14:00 UTC) but self-guards to only execute every 14 days.
    job_queue = app.job_queue
    job_queue.run_daily(
        scheduled_biweekly_broadcast,
        time=datetime.strptime("14:00", "%H:%M").time(),
        name="biweekly_scan",
    )
    job_queue.run_daily(
        scheduled_quarterly_market_trends,
        time=datetime.strptime("14:00", "%H:%M").time(),
        name="quarterly_market_trends",
    )

    if WEBHOOK_URL:
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=f"{WEBHOOK_URL}/webhook",
            url_path="webhook",
        )
    else:
        app.run_polling()


if __name__ == "__main__":
    main()
